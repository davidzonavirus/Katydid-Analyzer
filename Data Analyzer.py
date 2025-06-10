import sys
import os
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import scipy
import scipy.io
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout, QHBoxLayout, 
                           QWidget, QLabel, QFileDialog, QMessageBox, QFrame, QTableWidget, 
                           QTableWidgetItem, QSplitter, QTabWidget, QScrollArea, QSlider,
                           QDialog, QLineEdit, QGridLayout, QDoubleSpinBox, QCheckBox, QInputDialog,
                           QFormLayout, QDialogButtonBox)
from PyQt5.QtCore import Qt, QRectF, QPoint, QPropertyAnimation, QSize, pyqtSlot, QSequentialAnimationGroup, QEasingCurve, QPointF
from PyQt5.QtGui import QColor, QPalette, QFont, QDrag, QIcon, QLinearGradient, QRadialGradient, QPainter, QPen, QBrush, QPainterPath
from scipy.io import wavfile
import pandas as pd
from datetime import datetime
# Flag to track if Excel export is available
EXCEL_EXPORT_AVAILABLE = False
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_EXPORT_AVAILABLE = True
except ImportError:
    print("openpyxl not installed. Excel export will not be available.")


class AnimatedGradientWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)
        self.mouse_pos = QPoint(-100, -100)
        self.particles = []
        self.time = 0
        self.is_fullscreen = False
        
        # Initialize particles
        for _ in range(50):
            self.particles.append({
                'x': np.random.randint(0, 1000),
                'y': np.random.randint(0, 1000),
                'vx': np.random.randn() * 0.5,
                'vy': np.random.randn() * 0.5,
                'size': np.random.randint(3, 8)
            })
            
        # Start animation timer
        self.timer = self.startTimer(16)  # ~60 FPS
        
    def keyPressEvent(self, event):
        if event.key() == Qt.Key_F11:
            if self.is_fullscreen:
                self.window().showNormal()
            else:
                self.window().showFullScreen()
            self.is_fullscreen = not self.is_fullscreen
        super().keyPressEvent(event)
        
    def mouseMoveEvent(self, event):
        self.mouse_pos = event.pos()
        super().mouseMoveEvent(event)
        
    def timerEvent(self, event):
        self.time += 0.016
        
        # Update particle positions
        mouse_influence_radius = 150
        for p in self.particles:
            # Add mouse influence
            dx = self.mouse_pos.x() - p['x']
            dy = self.mouse_pos.y() - p['y']
            dist = np.sqrt(dx*dx + dy*dy)
            
            if dist < mouse_influence_radius:
                factor = (1 - dist/mouse_influence_radius) * 0.1
                p['vx'] += dx * factor
                p['vy'] += dy * factor
            
            # Update position
            p['x'] += p['vx']
            p['y'] += p['vy']
            
            # Add some random movement
            p['vx'] += (np.random.rand() - 0.5) * 0.1
            p['vy'] += (np.random.rand() - 0.5) * 0.1
            
            # Damping
            p['vx'] *= 0.99
            p['vy'] *= 0.99
            
            # Wrap around screen
            p['x'] = p['x'] % self.width()
            p['y'] = p['y'] % self.height()
        
        self.update()
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        # Create animated gradient background
        gradient = QLinearGradient(0, 0, self.width(), self.height())
        t = (np.sin(self.time * 0.5) + 1) * 0.5  # Oscillate between 0 and 1
        
        # Green to black gradient with animation
        gradient.setColorAt(0, QColor(0, int(100 + 55 * t), 0))  # Dark green
        gradient.setColorAt(1, QColor(0, 0, 0))  # Black
        painter.fillRect(self.rect(), gradient)
        
        # Draw connecting lines between nearby particles
        painter.setPen(QPen(QColor(0, 255, 0, 30), 1))
        max_dist = 100
        
        for i, p1 in enumerate(self.particles):
            for p2 in self.particles[i+1:]:
                dx = p1['x'] - p2['x']
                dy = p1['y'] - p2['y']
                dist = np.sqrt(dx*dx + dy*dy)
                
                if dist < max_dist:
                    alpha = int(255 * (1 - dist/max_dist) * 0.3)
                    painter.setPen(QPen(QColor(0, 255, 0, alpha), 1))
                    painter.drawLine(int(p1['x']), int(p1['y']), int(p2['x']), int(p2['y']))
        
        # Draw particles
        for p in self.particles:
            # Pulsating size based on time
            size_factor = 1 + 0.3 * np.sin(self.time * 2 + p['x'] * 0.01)
            size = p['size'] * size_factor
            
            # Create radial gradient for each particle
            particle_gradient = QRadialGradient(p['x'], p['y'], size * 2)
            particle_gradient.setColorAt(0, QColor(0, 255, 0, 150))
            particle_gradient.setColorAt(1, QColor(0, 255, 0, 0))
            
            painter.setBrush(QBrush(particle_gradient))
            painter.setPen(Qt.NoPen)
            painter.drawEllipse(QPointF(p['x'], p['y']), size, size)


class KatydidAnalyzer2(QMainWindow):
    def __init__(self):
        super().__init__()
        
        # Set window properties
        self.setWindowTitle("Katydid Analyzer 2")
        self.setGeometry(100, 100, 1200, 800)
        
        # Remove window frame and make fullscreen
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.showFullScreen()
        
        # Initialize data structures
        self.csv_data = None
        self.csv_file_path = None
        self.wav_data = None
        self.wav_file_path = None
        self.sample_rate = None
        self.pulses = []
        self.periods = []
        self.double_pulses = []
        self.double_pulse_sequences = []
        self.selected_peak = None
        self.peak_width = 1.0
        
        # Default copy settings (can be changed by user)
        self.copy_backward_ms = 200  # Default: 200ms before end marker
        self.copy_forward_ms = 300   # Default: 300ms after begin marker
        self.max_error_duration_ms = 10.0  # Default maximum error duration
        
        # Create status label for later use
        self.status_label = QLabel("Ready")
        self.status_label.setFont(QFont("Arial", 10))
        self.status_label.setStyleSheet("color: #666666;")
        
        # Set up the UI
        self.setup_start_screen()
        
        # Set up key event handling for the main window
        self.installEventFilter(self)
        
    def eventFilter(self, obj, event):
        # Handle key press events for the main window and other widgets
        try:
            if event.type() == event.KeyPress:
                # Check for Escape key to exit fullscreen or close the application
                if event.key() == Qt.Key_Escape:
                    self.close()
                    return True
                
                # Handle table-specific key presses
                if hasattr(self, 'table') and obj == self.table:
                    # Handle C key for Copy column functionality
                    if event.key() == Qt.Key_C:
                        print("C key pressed in table tab")
                        self.handle_copy_column()
                        return True
                    # Handle = key for saving waveform files
                    elif event.key() == Qt.Key_Equal:
                        print("= key pressed in table tab")
                        self.save_waveform_files()
                        return True
                
                # Handle tab-specific key presses
                if hasattr(self, 'tabs') and obj == self.tabs:
                    current_tab = self.tabs.currentIndex()
                    
                    # Handle key press events based on the current tab
                    if current_tab == 1:  # CSV table tab
                        if event.key() == Qt.Key_C:
                            print("C key pressed in table tab")
                            self.handle_copy_column()
                            return True
                        elif event.key() == Qt.Key_Equal:
                            print("= key pressed in table tab")
                            self.save_waveform_files()
                            return True
                    elif current_tab == 2:  # Period histogram tab
                        if event.key() == Qt.Key_K:
                            print("K key pressed in period histogram tab")
                            self.select_period_mode_range()
                            return True
                        elif event.key() == Qt.Key_L:
                            print("L key pressed in period histogram tab")
                            self.set_period_histogram_range()
                            return True
                    elif current_tab == 3:  # Ratio histogram tab
                        if event.key() == Qt.Key_K:
                            print("K key pressed in ratio histogram tab")
                            self.select_ratio_mode_range()
                            return True
                        elif event.key() == Qt.Key_L:
                            print("L key pressed in ratio histogram tab")
                            self.set_ratio_histogram_range()
                            return True
        except Exception as e:
            print(f"Error in eventFilter: {str(e)}")
        return super().eventFilter(obj, event)
        
    def setup_start_screen(self):
        # Create central widget with animated gradient background
        central_widget = AnimatedGradientWidget()
        self.setCentralWidget(central_widget)
        
        # Create layout for start screen
        layout = QVBoxLayout(central_widget)
        layout.setContentsMargins(50, 50, 50, 50)
        
        # Add title
        title_label = QLabel("Katydid Analyzer 2")
        title_label.setFont(QFont("Arial", 36, QFont.Bold))
        title_label.setStyleSheet("color: white; text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.5);")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # Add subtitle
        subtitle_label = QLabel("CSV and WAV Analysis Tool")
        subtitle_label.setFont(QFont("Arial", 18))
        subtitle_label.setStyleSheet("color: rgba(255, 255, 255, 0.8);")
        subtitle_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(subtitle_label)
        
        # Add spacer
        layout.addStretch(1)
        
        # Add load button
        load_button = QPushButton("Load Files")
        load_button.setFont(QFont("Arial", 14))
        load_button.setStyleSheet("""
            QPushButton {
                background-color: rgba(0, 128, 0, 0.7);
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background-color: rgba(0, 150, 0, 0.8);
            }
            QPushButton:pressed {
                background-color: rgba(0, 100, 0, 0.9);
            }
        """)
        load_button.clicked.connect(self.load_files)
        load_button.setMinimumWidth(200)
        load_button.setMinimumHeight(50)
        layout.addWidget(load_button, 0, Qt.AlignCenter)
        
        # Add spacer
        layout.addStretch(1)
        
        # Add exit button
        exit_button = QPushButton("Exit")
        exit_button.setFont(QFont("Arial", 12))
        exit_button.setStyleSheet("""
            QPushButton {
                background-color: rgba(150, 0, 0, 0.7);
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: rgba(200, 0, 0, 0.8);
            }
            QPushButton:pressed {
                background-color: rgba(100, 0, 0, 0.9);
            }
        """)
        exit_button.clicked.connect(self.close)
        exit_button.setMinimumWidth(100)
        layout.addWidget(exit_button, 0, Qt.AlignRight)
        
    def transition_to_analysis(self):
        # Prompt user to load CSV and WAV files
        self.load_files()
    
    def load_files(self):
        # Load CSV file
        if not self.load_csv_file():
            return
        
        # Load WAV file
        if not self.load_wav_file():
            return
        
        # Clear the start screen
        self.clear_start_screen()
        
        # Set up the analysis interface
        self.setup_analysis_interface()
    
    def clear_start_screen(self):
        # Remove all widgets from central widget
        for child in self.centralWidget().children():
            if isinstance(child, QWidget):
                child.deleteLater()
    
    def load_csv_file(self, file_path=None):
        # Open file dialog to select CSV file
        options = QFileDialog.Options()
        if file_path is None:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Open CSV File", "", "CSV Files (*.csv);;All Files (*)", options=options
            )
        
        if not file_path:
            return False
        
        try:
            # Load CSV file
            self.csv_data = pd.read_csv(file_path)
            self.csv_file_path = file_path
            
            # Process the CSV data to extract pulse information
            self.process_csv_data()
            
            # Update status
            self.status_label.setText(f"Loaded CSV file: {os.path.basename(file_path)}")
            
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load CSV file: {str(e)}")
            return False
    
    def process_csv_data(self):
        """Process the CSV data to extract pulse information"""
        # Reset pulse data
        self.periods = []
        
        # Check if CSV data is loaded
        if self.csv_data is None or self.csv_data.empty:
            return
        
        # Find the column names that match our required columns (case insensitive)
        period_col = next((col for col in self.csv_data.columns if col.lower() == 'period'), None)
        duration_col = next((col for col in self.csv_data.columns if col.lower() == 'duration (ms)' or col.lower() == 'duration'), None)
        ratio_col = next((col for col in self.csv_data.columns if col.lower() == 'pulse ratio'), None)
        amplitude_col = next((col for col in self.csv_data.columns if col.lower() == 'amplitude'), None)
        time_col = next((col for col in self.csv_data.columns if col.lower() == 'time (ms)' or col.lower() == 'time'), None)
        
        # Check if we have the required columns
        if not (duration_col and ratio_col):
            QMessageBox.warning(self, "Warning", "CSV file does not contain required columns (Duration, Pulse Ratio)")
            return
        
        # Extract pulse information
        for i, (_, row) in enumerate(self.csv_data.iterrows()):
            period = {}
            
            # Add period number if available
            if period_col and pd.notna(row[period_col]):
                period['period'] = row[period_col]
            else:
                period['period'] = i + 1
            
            # Add duration
            if pd.notna(row[duration_col]):
                period['duration'] = row[duration_col]
            
            # Add pulse ratio
            if pd.notna(row[ratio_col]):
                period['ratio'] = row[ratio_col]
            
            # Add amplitude if available
            if amplitude_col and pd.notna(row[amplitude_col]):
                period['amplitude'] = row[amplitude_col]
            
            # Add time if available
            if time_col and pd.notna(row[time_col]):
                period['time'] = row[time_col]
            
            # Add to periods list
            self.periods.append(period)
        
        # Extract pulse information for waveform display
        self.pulses = []
        for period in self.periods:
            if 'time' in period and 'amplitude' in period:
                self.pulses.append({
                    'time': period['time'],
                    'amplitude': period['amplitude']
                })
        
        print(f"Processed {len(self.periods)} periods from CSV data")
    
    def load_wav_file(self):
        # Open file dialog to select WAV file
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open WAV File", "", "WAV Files (*.wav);;All Files (*)", options=options
        )
        
        if not file_path:
            return False
        
        try:
            # Load WAV file
            self.sample_rate, self.wav_data = wavfile.read(file_path)
            self.wav_file_path = file_path
            
            # Convert to mono if stereo
            if len(self.wav_data.shape) > 1 and self.wav_data.shape[1] > 1:
                self.wav_data = np.mean(self.wav_data, axis=1)
            
            # Normalize the data
            self.wav_data = self.wav_data.astype(np.float32)
            if np.max(np.abs(self.wav_data)) > 0:
                self.wav_data = self.wav_data / np.max(np.abs(self.wav_data))
            
            # Update status
            self.status_label.setText(f"Loaded WAV file: {os.path.basename(file_path)}")
            
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load WAV file: {str(e)}")
            return False
    
    def setup_analysis_interface(self):
        # Clear the start screen first
        self.clear_start_screen()
        
        # Create a new central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Add header with exit button only (no title)
        header_layout = QHBoxLayout()
        
        # Add spacer to push exit button to the right
        header_layout.addStretch()
        
        # Add exit button
        exit_button = QPushButton("X")
        exit_button.setFixedSize(30, 30)
        exit_button.clicked.connect(self.close)
        exit_button.setStyleSheet("""
            QPushButton {
                color: white;
                background-color: rgba(150, 0, 0, 0.7);
                font-size: 16px;
                font-weight: bold;
                border: none;
                border-radius: 15px;
            }
            QPushButton:hover {
                background-color: rgba(200, 0, 0, 0.8);
            }
            QPushButton:pressed {
                background-color: rgba(100, 0, 0, 0.9);
            }
        """)
        header_layout.addWidget(exit_button)
        
        # Add header to main layout
        main_layout.addLayout(header_layout)
        
        # Create tab widget
        self.tabs = QTabWidget()
        
        # Install event filter on the tab widget to catch key events
        self.tabs.installEventFilter(self)
        
        # Create the initial three tabs
        self.create_csv_table_tab()
        self.create_waveform_tab()
        self.create_period_histogram_tab()
        
        # Ratio histogram tab will be created after period range is selected
        
        # Add tab widget to main layout
        main_layout.addWidget(self.tabs)
        
        # Add status bar
        self.status_label = QLabel("Ready")
        self.status_label.setFont(QFont("Arial", 10))
        self.status_label.setStyleSheet("color: #666666;")
        main_layout.addWidget(self.status_label)
        
        # Set keyboard focus to the current tab
        self.tabs.currentWidget().setFocus()
    
    def create_csv_table_tab(self):
        # Create tab widget
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Create table
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(["Pulse", "Amplitude", "Time", "TimeB", "Period Duration", 
                                              "Ratio", "ex & in", "Sequencing", "Copy"])
        
        # Make the table read-only to prevent user edits
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        # Populate table with data
        self.table.setRowCount(len(self.csv_data))
        
        for row, (_, csv_row) in enumerate(self.csv_data.iterrows()):
            # Find the column names that match our required columns (case insensitive)
            # For Pulse column, we'll just use the row number (1-based)
            amplitude_col = next((col for col in self.csv_data.columns if col.lower() == 'amplitude'), None)
            time_col = next((col for col in self.csv_data.columns if col.lower() == 'time (ms)' or col.lower() == 'time'), None)
            period_duration_col = next((col for col in self.csv_data.columns if col.lower() == 'period duration' or col.lower() == 'duration (ms)' or col.lower() == 'duration'), None)
            ratio_col = next((col for col in self.csv_data.columns if col.lower() == 'ratio' or col.lower() == 'pulse ratio'), None)
            
            # Add pulse number (1-based) to first column
            pulse_item = QTableWidgetItem(f"{row+1}")
            self.table.setItem(row, 0, pulse_item)
            
            # Set data for each column (starting from column 1 since we already set column 0 with pulse number)
            for col, (header, csv_col) in enumerate(zip(
                ["Amplitude", "Time", "TimeB", "Period Duration", "Ratio"],
                [amplitude_col, time_col, None, period_duration_col, ratio_col]
            ), start=1):
                if header == "TimeB":
                    # TimeB will be calculated after all rows are populated
                    item = QTableWidgetItem("")
                elif csv_col and pd.notna(csv_row[csv_col]):
                    value = csv_row[csv_col]
                    if isinstance(value, (int, float)):
                        if header in ["Time", "TimeB", "Period Duration"]:
                            item = QTableWidgetItem(f"{value:.2f}")
                        elif header == "Ratio":
                            item = QTableWidgetItem(f"{value:.4f}")
                        else:
                            item = QTableWidgetItem(f"{value}")
                    else:
                        item = QTableWidgetItem(str(value))
                else:
                    item = QTableWidgetItem("")
                
                self.table.setItem(row, col, item)
        
        # Calculate TimeB values (time between pulses)
        for row in range(1, self.table.rowCount()):
            time_item_prev = self.table.item(row-1, 2)  # Previous Time column
            time_item_curr = self.table.item(row, 2)    # Current Time column
            
            if time_item_prev and time_item_curr and time_item_prev.text() and time_item_curr.text():
                try:
                    time_prev = float(time_item_prev.text())
                    time_curr = float(time_item_curr.text())
                    time_between = time_curr - time_prev
                    self.table.setItem(row, 3, QTableWidgetItem(f"{time_between:.2f}"))
                except ValueError:
                    self.table.setItem(row, 3, QTableWidgetItem(""))
            else:
                self.table.setItem(row, 3, QTableWidgetItem(""))
                
        # Add empty cells for the new columns
        for row in range(self.table.rowCount()):
            for col in range(6, 9):
                self.table.setItem(row, col, QTableWidgetItem(""))
                
        # Set up key event handling for the table
        self.table.installEventFilter(self)
        
        # Resize columns to content
        self.table.resizeColumnsToContents()
        
        # Make columns 1.25x bigger
        for col in range(self.table.columnCount()):
            width = self.table.columnWidth(col)
            self.table.setColumnWidth(col, int(width * 1.25))
        
        # Add table to layout
        layout.addWidget(self.table)
        
        # Add the tab to the tab widget
        self.tabs.addTab(tab, "CSV Table")
    
    def create_waveform_tab(self):
        # Create tab widget
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Create matplotlib figure
        self.waveform_fig = Figure(figsize=(10, 6), tight_layout=True)
        self.waveform_canvas = FigureCanvas(self.waveform_fig)
        self.waveform_ax = self.waveform_fig.add_subplot(111)
        
        # Plot waveform
        self.update_waveform_plot()
        
        # Add canvas to layout
        layout.addWidget(self.waveform_canvas)
        
        # Add navigation instructions
        nav_label = QLabel("Navigation: W/S - Zoom In/Out, A/D - Move Left/Right")
        nav_label.setFont(QFont("Arial", 9))
        nav_label.setStyleSheet("color: #666666;")
        layout.addWidget(nav_label)
        
        # Connect key press event
        self.waveform_canvas.setFocusPolicy(Qt.StrongFocus)
        self.waveform_canvas.mpl_connect('key_press_event', self.on_waveform_key_press)
        
        # Initialize view limits
        self.waveform_view_limits = {'xmin': 0, 'xmax': 1000, 'zoom_factor': 1.0}
        
        # Add the tab to the tab widget
        self.tabs.addTab(tab, "Waveform")
    
    def on_waveform_key_press(self, event):
        # Handle key press events for waveform navigation
        if not hasattr(self, 'waveform_view_limits'):
            self.waveform_view_limits = {'xmin': 0, 'xmax': 1000, 'zoom_factor': 1.0}
        
        # Get current x limits
        xmin, xmax = self.waveform_ax.get_xlim()
        view_width = xmax - xmin
        
        # Handle key presses
        if event.key == 'w':  # Zoom in
            zoom_factor = 0.8
            center = (xmin + xmax) / 2
            new_width = view_width * zoom_factor
            self.waveform_ax.set_xlim(center - new_width/2, center + new_width/2)
            self.waveform_view_limits['zoom_factor'] *= zoom_factor
        elif event.key == 's':  # Zoom out
            zoom_factor = 1.25
            center = (xmin + xmax) / 2
            new_width = view_width * zoom_factor
            self.waveform_ax.set_xlim(center - new_width/2, center + new_width/2)
            self.waveform_view_limits['zoom_factor'] *= zoom_factor
        elif event.key == 'a':  # Move left
            move_amount = view_width * 0.2
            self.waveform_ax.set_xlim(xmin - move_amount, xmax - move_amount)
        elif event.key == 'd':  # Move right
            move_amount = view_width * 0.2
            self.waveform_ax.set_xlim(xmin + move_amount, xmax + move_amount)
        
        # Update the canvas
        self.waveform_canvas.draw()
        
        # Update status
        self.status_label.setText(f"View: {xmin:.1f} - {xmax:.1f} ms, Zoom: {1/self.waveform_view_limits['zoom_factor']:.1f}x")
    
    def update_waveform_plot(self):
        # Clear the plot
        self.waveform_ax.clear()
        
        # Plot the waveform
        if self.wav_data is not None and self.sample_rate is not None:
            # Create time array in milliseconds
            time_ms = np.arange(len(self.wav_data)) / self.sample_rate * 1000
            
            # Plot the waveform
            self.waveform_ax.plot(time_ms, self.wav_data, 'k-', linewidth=0.5, label='Waveform')
            
            # Plot pulses if available
            if hasattr(self, 'pulses') and self.pulses:
                for pulse in self.pulses:
                    if 'time' in pulse and 'amplitude' in pulse:
                        self.waveform_ax.plot(pulse['time'], pulse['amplitude'], 'ro', markersize=5)
            
            # Mark double pulses if identified
            if hasattr(self, 'double_pulses') and self.double_pulses:
                for pulse in self.double_pulses:
                    if 'time' in pulse and 'amplitude' in pulse:
                        self.waveform_ax.plot(pulse['time'], pulse['amplitude'], 'bo', markersize=5)
            
            # Set labels and title
            self.waveform_ax.set_xlabel('Time (ms)')
            self.waveform_ax.set_ylabel('Amplitude')
            self.waveform_ax.set_title('Waveform with Pulses')
            
            # Add legend
            self.waveform_ax.legend(loc='upper right')
            
            # Set grid
            self.waveform_ax.grid(True, linestyle='--', alpha=0.7)
            
            # Update canvas
            self.waveform_canvas.draw()
    
    def create_period_histogram_tab(self):
        # Create tab widget
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Create matplotlib figure
        self.period_fig = Figure(figsize=(10, 5), tight_layout=True)
        self.period_canvas = FigureCanvas(self.period_fig)
        self.period_ax = self.period_fig.add_subplot(111)
        
        # Plot period histogram
        self.update_period_histogram()
        
        # Add canvas to layout
        layout.addWidget(self.period_canvas)
        
        # Add navigation instructions
        nav_label = QLabel("Navigation: W/S - Zoom In/Out, A/D - Move Left/Right")
        nav_label.setFont(QFont("Arial", 9))
        nav_label.setStyleSheet("color: #666666;")
        layout.addWidget(nav_label)
        
        # Add info label for clicked bars
        self.period_info_label = QLabel("Click on a bar to see details")
        self.period_info_label.setFont(QFont("Arial", 9))
        self.period_info_label.setStyleSheet("color: #666666; background-color: #f0f0f0; padding: 5px; border-radius: 3px;")
        layout.addWidget(self.period_info_label)
        
        # Connect key press event
        self.period_canvas.setFocusPolicy(Qt.StrongFocus)
        self.period_canvas.mpl_connect('key_press_event', self.on_period_histogram_key_press)
        self.period_canvas.mpl_connect('button_press_event', self.on_period_histogram_click)
        
        # Initialize view limits
        self.period_view_limits = {'xmin': 0, 'xmax': 100, 'zoom_factor': 1.0}
        
        # Add the tab to the tab widget
        self.tabs.addTab(tab, "Period Histogram")
    
    def on_period_histogram_click(self, event):
        # Handle click events on the period histogram
        if event.inaxes != self.period_ax:
            return
        
        # Get the x value (period duration) where the user clicked
        clicked_x = event.xdata
        
        # Find periods close to the clicked value
        if hasattr(self, 'periods') and self.periods:
            durations = [p['duration'] for p in self.periods if 'duration' in p]
            if durations:
                # Find the closest duration to the clicked value
                closest_idx = np.argmin(np.abs(np.array(durations) - clicked_x))
                closest_duration = durations[closest_idx]
                
                # Get the corresponding period data
                period_data = self.periods[closest_idx]
                
                # Update the info label with details
                info_text = f"Period: {period_data.get('period', 'N/A')}, Duration: {period_data.get('duration', 'N/A'):.2f} ms, "
                info_text += f"Ratio: {period_data.get('ratio', 'N/A'):.4f}, "
                info_text += f"Time: {period_data.get('time', 'N/A'):.2f} ms"
                
                self.period_info_label.setText(info_text)
    
    def on_period_histogram_key_press(self, event):
        # Handle key press events for period histogram navigation
        if not hasattr(self, 'period_view_limits'):
            self.period_view_limits = {'xmin': 0, 'xmax': 100, 'zoom_factor': 1.0}
        
        # Get current x limits
        xmin, xmax = self.period_ax.get_xlim()
        view_width = xmax - xmin
        
        # Handle key presses
        if event.key == 'w':  # Zoom in
            zoom_factor = 0.8
            center = (xmin + xmax) / 2
            new_width = view_width * zoom_factor
            self.period_ax.set_xlim(center - new_width/2, center + new_width/2)
            self.period_view_limits['zoom_factor'] *= zoom_factor
        elif event.key == 's':  # Zoom out
            zoom_factor = 1.25
            center = (xmin + xmax) / 2
            new_width = view_width * zoom_factor
            self.period_ax.set_xlim(center - new_width/2, center + new_width/2)
            self.period_view_limits['zoom_factor'] *= zoom_factor
        elif event.key == 'a':  # Move left
            move_amount = view_width * 0.2
            self.period_ax.set_xlim(xmin - move_amount, xmax - move_amount)
        elif event.key == 'd':  # Move right
            move_amount = view_width * 0.2
            self.period_ax.set_xlim(xmin + move_amount, xmax + move_amount)
        elif event.key == 'k':  # Select mode peak and set range
            self.select_period_mode_range()
        
        # Update the canvas
        self.period_canvas.draw()
        
        # Update status
        self.status_label.setText(f"Period view: {xmin:.1f} - {xmax:.1f} ms, Zoom: {1/self.period_view_limits['zoom_factor']:.1f}x")
    
    def select_period_mode_range(self):
        # Select mode peak and set range for period histogram
        if hasattr(self, 'period_hist_data') and self.period_hist_data:
            mode_value = self.period_hist_data['mode_value']
            left_bar = self.period_hist_data['left_bar']
            right_bar = self.period_hist_data['right_bar']
            recommended_min = self.period_hist_data['recommended_min']
            recommended_max = self.period_hist_data['recommended_max']
            
            # Create a simple dialog asking if the variation is OK
            message = f"Mode peak detected at {mode_value:.2f} ms\n"
            message += f"Closest bars at {left_bar:.2f} ms and {right_bar:.2f} ms\n"
            message += f"Recommended range: {recommended_min:.2f} - {recommended_max:.2f} ms\n\n"
            message += "Is this variation acceptable?"
            
            reply = QMessageBox.question(self, "Confirm Period Variation", message,
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            
            if reply == QMessageBox.No:
                # User wants to adjust the variation
                dialog = QDialog(self)
                dialog.setWindowTitle("Adjust Period Variation")
                # Make dialog non-modal so user can interact with main window
                dialog.setWindowModality(Qt.NonModal)
                layout = QVBoxLayout(dialog)
                
                # Add explanation label
                layout.addWidget(QLabel(f"Mode value: {mode_value:.2f} ms"))
                layout.addWidget(QLabel("Set custom values for left and right sides:"))
                
                # Add variation inputs for both left and right sides
                variation_layout = QGridLayout()
                
                # Left/Right label for clarity
                variation_layout.addWidget(QLabel("Left/Right"), 0, 0)
                
                # Left value (absolute)
                variation_layout.addWidget(QLabel("Left value (ms):"), 1, 0)
                left_value_input = QDoubleSpinBox()
                left_value_input.setRange(0, mode_value)
                left_value_input.setValue(recommended_min)
                left_value_input.setDecimals(2)
                left_value_input.setSingleStep(0.1)
                variation_layout.addWidget(left_value_input, 1, 1)
                
                # Right value (absolute)
                variation_layout.addWidget(QLabel("Right value (ms):"), 2, 0)
                right_value_input = QDoubleSpinBox()
                right_value_input.setRange(mode_value, 50)
                right_value_input.setValue(recommended_max)
                right_value_input.setDecimals(2)
                right_value_input.setSingleStep(0.1)
                variation_layout.addWidget(right_value_input, 2, 1)
                
                # Preview of resulting range
                preview_label = QLabel(f"Resulting range: {left_value_input.value():.2f} - {right_value_input.value():.2f} ms")
                
                # Create a figure for real-time histogram preview
                preview_figure = Figure(figsize=(6, 3), dpi=100)
                preview_canvas = FigureCanvas(preview_figure)
                preview_ax = preview_figure.add_subplot(111)
                
                # Function to update the preview histogram and the main histogram
                def update_preview_histogram():
                    # Clear the axis
                    preview_ax.clear()
                    
                    # Get current values
                    range_min = left_value_input.value()
                    range_max = right_value_input.value()
                    
                    # Update text preview
                    preview_label.setText(f"Resulting range: {range_min:.2f} - {range_max:.2f} ms")
                    
                    # Also update the main histogram in real-time
                    # Store current axis limits before clearing
                    current_xlim = self.period_ax.get_xlim() if hasattr(self.period_ax, 'get_xlim') else (0, 50)
                    current_ylim = self.period_ax.get_ylim() if hasattr(self.period_ax, 'get_ylim') else None
                    
                    # Clear the main period histogram
                    self.period_ax.clear()
                    
                    # Redraw the main histogram
                    if hasattr(self, 'periods') and self.periods:
                        durations = [p['duration'] for p in self.periods if 'duration' in p]
                        if durations and hasattr(self, 'period_bins'):
                            # Plot histogram with bars that touch each other (no gaps)
                            self.period_ax.hist(durations, bins=self.period_bins, alpha=0.7, 
                                              color='green', edgecolor='black', linewidth=0.5,
                                              align='left', rwidth=1.0)
                            
                            # Mark the mode with a vertical line
                            self.period_ax.axvline(x=mode_value, color='red', linestyle='-', linewidth=2, 
                                                 label=f'Mode: {mode_value:.2f} ms')
                            
                            # Show the current range being adjusted
                            self.period_ax.axvline(x=range_min, color='blue', linestyle='--', linewidth=1.5, 
                                                 label=f'Min: {range_min:.2f} ms')
                            self.period_ax.axvline(x=range_max, color='blue', linestyle='--', linewidth=1.5, 
                                                 label=f'Max: {range_max:.2f} ms')
                            
                            # Highlight the selected range
                            self.period_ax.axvspan(range_min, range_max, alpha=0.3, color='green',
                                                label=f'Range: {range_min:.2f}-{range_max:.2f} ms')
                            
                            # Set labels and title
                            self.period_ax.set_xlabel('Period Duration (ms)', fontsize=10, fontweight='bold')
                            self.period_ax.set_ylabel('Frequency', fontsize=10, fontweight='bold')
                            self.period_ax.set_title('Distribution of Period Durations', fontsize=12, fontweight='bold')
                            
                            # Set grid
                            self.period_ax.grid(True, linestyle='--', alpha=0.7)
                            
                            # Make tick labels more visible
                            self.period_ax.tick_params(axis='both', which='major', labelsize=9)
                            
                            # Add a legend
                            self.period_ax.legend(loc='upper right', fontsize=8)
                            
                            # Restore previous axis limits to maintain zoom level
                            self.period_ax.set_xlim(current_xlim)
                            if current_ylim is not None:
                                self.period_ax.set_ylim(current_ylim)
                            
                            # Update the main canvas
                            self.period_canvas.draw()
                    
                    # Plot the histogram using stored data
                    if hasattr(self, 'period_hist_data'):
                        hist_data = self.period_hist_data
                        bins = hist_data['bin_edges']
                        hist = hist_data['hist']
                        
                        # Plot the histogram
                        bin_centers = [(bins[i] + bins[i+1])/2 for i in range(len(bins)-1)]
                        bin_width = bins[1] - bins[0]
                        preview_ax.bar(bin_centers, hist, width=bin_width, alpha=0.7, 
                                      color='green', edgecolor='black', linewidth=0.5)
                        
                        # Mark the mode
                        preview_ax.axvline(x=mode_value, color='red', linestyle='-', linewidth=2)
                        
                        # Highlight the selected range
                        preview_ax.axvspan(range_min, range_max, alpha=0.3, color='blue')
                        
                        # Set labels and limits
                        preview_ax.set_xlabel('Period Duration (ms)')
                        preview_ax.set_ylabel('Frequency')
                        preview_ax.set_xlim(0, 50)  # Standard x-axis limits for period histogram
                        preview_ax.grid(True, linestyle='--', alpha=0.7)
                    
                    # Redraw the canvas
                    preview_canvas.draw()
                
                # Connect value changes to update preview
                left_value_input.valueChanged.connect(update_preview_histogram)
                right_value_input.valueChanged.connect(update_preview_histogram)
                
                # Initial update
                update_preview_histogram()
                
                layout.addLayout(variation_layout)
                layout.addWidget(preview_label)
                
                # Add buttons
                buttons = QHBoxLayout()
                ok_button = QPushButton("OK")
                ok_button.clicked.connect(dialog.accept)
                cancel_button = QPushButton("Cancel")
                cancel_button.clicked.connect(dialog.reject)
                buttons.addWidget(ok_button)
                buttons.addWidget(cancel_button)
                layout.addLayout(buttons)
                
                if dialog.exec_() == QDialog.Accepted:
                    # Get the custom values directly
                    range_min = left_value_input.value()
                    range_max = right_value_input.value()
                else:
                    # User cancelled the custom variation dialog, use recommended values
                    range_min = recommended_min
                    range_max = recommended_max
            else:
                # User accepted the recommended variation
                range_min = recommended_min
                range_max = recommended_max
            
            # Store the period range
            self.period_range = (range_min, range_max)
            
            # Clear previous range visualization if any
            self.update_period_histogram()
            
            # Highlight the selected range on the histogram
            self.period_ax.axvspan(range_min, range_max, alpha=0.3, color='green', 
                                   label=f'Selected: {range_min:.2f}-{range_max:.2f} ms')
            
            # Update the canvas
            self.period_canvas.draw()
            
            # Update status
            self.status_label.setText(f"Selected period range: {range_min:.2f} - {range_max:.2f} ms")
            
            # Create the ratio histogram tab if it doesn't exist yet
            if self.tabs.count() < 4:  # If we don't have a ratio tab yet
                self.create_ratio_histogram_tab()
            
            # Now update the ratio histogram based on the selected period range
            self.update_ratio_histogram()
            
            # Switch to the ratio histogram tab (it should be the last tab)
            self.tabs.setCurrentIndex(self.tabs.count() - 1)
            
            # Inform the user about the next step
            QMessageBox.information(self, "Next Step", 
                                    "Now press 'K' in the Ratio Histogram tab to select the ratio range.")
    
    def on_period_histogram_key_press(self, event):
        # Handle key press events for period histogram navigation
        if not hasattr(self, 'period_view_limits'):
            self.period_view_limits = {'xmin': 0, 'xmax': 50, 'zoom_factor': 1.0}
        
        # Get current x limits
        xmin, xmax = self.period_ax.get_xlim()
        view_width = xmax - xmin
        
        # Handle key presses
        if event.key == 'w':  # Zoom in
            zoom_factor = 0.8
            center = (xmin + xmax) / 2
            new_width = view_width * zoom_factor
            self.period_ax.set_xlim(center - new_width/2, center + new_width/2)
            self.period_view_limits['zoom_factor'] *= zoom_factor
        elif event.key == 's':  # Zoom out
            zoom_factor = 1.25
            center = (xmin + xmax) / 2
            new_width = view_width * zoom_factor
            self.period_ax.set_xlim(center - new_width/2, center + new_width/2)
            self.period_view_limits['zoom_factor'] *= zoom_factor
        elif event.key == 'a':  # Move left
            move_amount = view_width * 0.2
            self.period_ax.set_xlim(xmin - move_amount, xmax - move_amount)
        elif event.key == 'd':  # Move right
            move_amount = view_width * 0.2
            self.period_ax.set_xlim(xmin + move_amount, xmax + move_amount)
        elif event.key == 'k':  # Select mode peak and set range
            self.select_period_mode_range()
        elif event.key == 'l':  # Manually set x-axis range
            self.set_period_histogram_range()
        
        # Update the canvas
        self.period_canvas.draw()
        
        # Update status
        self.status_label.setText(f"Period view: {xmin:.1f} - {xmax:.1f} ms, Zoom: {1/self.period_view_limits['zoom_factor']:.1f}x")
    
    def set_period_histogram_range(self):
        # Allow user to manually set the x-axis range for period histogram
        current_min, current_max = self.period_ax.get_xlim()
        
        # Create dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Set Period Histogram X-Axis Range")
        layout = QVBoxLayout(dialog)
        
        # Add explanation
        layout.addWidget(QLabel("Set the x-axis range for the period histogram:"))
        
        # Add min/max inputs
        range_layout = QGridLayout()
        range_layout.addWidget(QLabel("Minimum (ms):"), 0, 0)
        min_input = QDoubleSpinBox()
        min_input.setRange(0, 100)
        min_input.setValue(current_min)
        min_input.setDecimals(2)
        min_input.setSingleStep(1.0)
        range_layout.addWidget(min_input, 0, 1)
        
        range_layout.addWidget(QLabel("Maximum (ms):"), 1, 0)
        max_input = QDoubleSpinBox()
        max_input.setRange(0, 100)
        max_input.setValue(current_max)
        max_input.setDecimals(2)
        max_input.setSingleStep(1.0)
        range_layout.addWidget(max_input, 1, 1)
        
        layout.addLayout(range_layout)
        
        # Add buttons
        buttons = QHBoxLayout()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(dialog.accept)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(dialog.reject)
        buttons.addWidget(ok_button)
        buttons.addWidget(cancel_button)
        layout.addLayout(buttons)
        
        if dialog.exec_() == QDialog.Accepted:
            # Get the custom range values
            new_min = min_input.value()
            new_max = max_input.value()
            
            # Validate range
            if new_min >= new_max:
                QMessageBox.warning(self, "Invalid Range", "Minimum must be less than maximum.")
                return
            
            # Set the new range
            self.period_ax.set_xlim(new_min, new_max)
            self.period_canvas.draw()
            
            # Update status
            self.status_label.setText(f"Period view range set to: {new_min:.1f} - {new_max:.1f} ms")
    
        
        
        
        
        
        
        
        
            
    def update_table_with_ranges(self):
        # Check if we have both period and ratio ranges set
        if not hasattr(self, 'period_range') or not hasattr(self, 'ratio_range'):
            QMessageBox.warning(self, "Warning", "Both period and ratio ranges must be set first")
            return
        
        period_min, period_max = self.period_range
        ratio_min, ratio_max = self.ratio_range
        
        # Make sure we have a table to update
        if not hasattr(self, 'table'):
            return
        
        # Update the table with pulse information
        for i, period in enumerate(self.periods):
            if i < self.table.rowCount():
                # Add pulse number
                pulse_item = QTableWidgetItem(f"{i+1}")
                self.table.setItem(i, 0, pulse_item)
                
                # Add amplitude if available
                if 'amplitude' in period:
                    amp_item = QTableWidgetItem(f"{period['amplitude']:.2f}")
                    self.table.setItem(i, 1, amp_item)
                
                # Add time if available
                if 'time' in period:
                    time_item = QTableWidgetItem(f"{period['time']:.2f}")
                    self.table.setItem(i, 2, time_item)
                
                # Calculate TimeB (time between pulses)
                if i > 0 and 'time' in period and 'time' in self.periods[i-1]:
                    time_prev = self.periods[i-1]['time']
                    time_curr = period['time']
                    time_between = time_curr - time_prev
                    timeb_item = QTableWidgetItem(f"{time_between:.2f}")
                    self.table.setItem(i, 3, timeb_item)
                
                # Add period duration if available
                if 'duration' in period:
                    duration_item = QTableWidgetItem(f"{period['duration']:.2f}")
                    self.table.setItem(i, 4, duration_item)
                
                # Add ratio if available
                if 'ratio' in period:
                    ratio_item = QTableWidgetItem(f"{period['ratio']:.4f}")
                    self.table.setItem(i, 5, ratio_item)
                
                # Check if within both ranges
                duration_in_range = period_min <= period.get('duration', 0) <= period_max if 'duration' in period else False
                ratio_in_range = ratio_min <= period.get('ratio', 0) <= ratio_max if 'ratio' in period else False
                within_range = duration_in_range and ratio_in_range
                
                # Add ex & in indicator - now at column 6 with TimeB added
                if within_range:
                    if period.get('ratio', 0) < 0.5:
                        # External pulse
                        ex_in_item = QTableWidgetItem('ex')
                        period['is_valid_short'] = True
                    elif i > 0 and self.periods[i-1].get('is_valid_short', False):
                        # Internal pulse
                        ex_in_item = QTableWidgetItem('in')
                    else:
                        # Invalid pulse
                        ex_in_item = QTableWidgetItem('z')
                        period['is_valid_short'] = False
                else:
                    # Invalid pulse
                    ex_in_item = QTableWidgetItem('z')
                    period['is_valid_short'] = False
                
                self.table.setItem(i, 6, ex_in_item)  # Column 6 for ex & in with TimeB added
                
                # Clear sequencing column
                self.table.setItem(i, 7, QTableWidgetItem(""))
                
                # Highlight row if outside range (light red background)
                for col in range(self.table.columnCount()):
                    item = self.table.item(i, col)
                    if item:
                        if not within_range:
                            item.setBackground(QColor(255, 200, 200))  # Light red
                        else:
                            item.setBackground(QColor(255, 255, 255))  # White
        
        # Resize columns to content
        self.table.resizeColumnsToContents()
        
        # Update status
        self.status_label.setText(f"Table updated with ranges: Period {period_min:.2f}-{period_max:.2f} ms, Ratio {ratio_min:.3f}-{ratio_max:.3f}")
    
    def select_ratio_mode_range(self):
        # Select mode peak and set range for ratio histogram
        if hasattr(self, 'ratio_hist_data') and self.ratio_hist_data:
            mode_value = self.ratio_hist_data['mode_value']
            left_bar = self.ratio_hist_data['left_bar']
            right_bar = self.ratio_hist_data['right_bar']
            recommended_min = self.ratio_hist_data['recommended_min']
            recommended_max = self.ratio_hist_data['recommended_max']
            
            # Create a simple dialog asking if the variation is OK
            message = f"Mode peak detected at {mode_value:.3f}\n"
            message += f"Closest bars at {left_bar:.3f} and {right_bar:.3f}\n"
            message += f"Recommended range: {recommended_min:.3f} - {recommended_max:.3f}\n\n"
            message += "Is this variation acceptable?"
            
            reply = QMessageBox.question(self, "Confirm Ratio Variation", message,
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            
            if reply == QMessageBox.No:
                # User wants to adjust the variation
                dialog = QDialog(self)
                dialog.setWindowTitle("Adjust Ratio Variation")
                # Make dialog non-modal so user can interact with main window
                dialog.setWindowModality(Qt.NonModal)
                layout = QVBoxLayout(dialog)
                
                # Add explanation label
                layout.addWidget(QLabel(f"Mode value: {mode_value:.3f}"))
                layout.addWidget(QLabel("Set custom variation for left and right sides:"))
                
                # Add variation inputs for both left and right sides
                variation_layout = QGridLayout()
                
                # Left variation
                variation_layout.addWidget(QLabel("Left variation:"), 0, 0)
                left_variation_input = QDoubleSpinBox()
                left_variation_input.setRange(0.001, 0.5)
                left_variation_input.setValue(mode_value - recommended_min)
                left_variation_input.setDecimals(3)
                left_variation_input.setSingleStep(0.005)
                variation_layout.addWidget(left_variation_input, 0, 1)
                
                # Right variation
                variation_layout.addWidget(QLabel("Right variation:"), 1, 0)
                right_variation_input = QDoubleSpinBox()
                right_variation_input.setRange(0.001, 0.5)
                right_variation_input.setValue(recommended_max - mode_value)
                right_variation_input.setDecimals(3)
                right_variation_input.setSingleStep(0.005)
                variation_layout.addWidget(right_variation_input, 1, 1)
                
                # Preview of resulting range
                preview_label = QLabel(f"Resulting range: {max(0, mode_value - left_variation_input.value()):.3f} - {min(1, mode_value + right_variation_input.value()):.3f}")
                
                # Create a figure for real-time histogram preview
                preview_figure = Figure(figsize=(6, 3), dpi=100)
                preview_canvas = FigureCanvas(preview_figure)
                preview_ax = preview_figure.add_subplot(111)
                
                # Function to update both the preview and the main histogram
                def update_preview():
                    # Get current variation values
                    left_var = left_variation_input.value()
                    right_var = right_variation_input.value()
                    
                    # Calculate range
                    range_min = max(0, mode_value - left_var)
                    range_max = min(1, mode_value + right_var)
                    
                    # Update text preview
                    preview_label.setText(f"Resulting range: {range_min:.3f} - {range_max:.3f}")
                    
                    # Clear the preview axis
                    preview_ax.clear()
                    
                    # Also update the main histogram in real-time
                    # Store current axis limits before clearing
                    current_xlim = self.ratio_ax.get_xlim() if hasattr(self.ratio_ax, 'get_xlim') else (0, 1)
                    current_ylim = self.ratio_ax.get_ylim() if hasattr(self.ratio_ax, 'get_ylim') else None
                    
                    # Clear the main ratio histogram
                    self.ratio_ax.clear()
                    
                    # Redraw the main histogram
                    if hasattr(self, 'ratio_hist_data'):
                        hist_data = self.ratio_hist_data
                        bin_centers = hist_data['bin_centers']
                        hist = hist_data['hist']
                        bin_width = hist_data['bin_edges'][1] - hist_data['bin_edges'][0]
                        
                        # Plot both histograms (preview and main)
                        # Preview histogram
                        preview_ax.clear()
                        preview_ax.bar(bin_centers, hist, width=bin_width, alpha=0.7, 
                                      color='blue', edgecolor='black', linewidth=0.5)
                        preview_ax.axvline(x=mode_value, color='red', linestyle='-', linewidth=2)
                        preview_ax.axvline(x=range_min, color='blue', linestyle='--', linewidth=1.5)
                        preview_ax.axvline(x=range_max, color='blue', linestyle='--', linewidth=1.5)
                        preview_ax.axvspan(range_min, range_max, alpha=0.2, color='green')
                        preview_ax.set_xlim(0, 1)
                        preview_ax.set_xlabel('Ratio')
                        preview_ax.set_ylabel('Frequency')
                        
                        # Main histogram
                        self.ratio_ax.bar(bin_centers, hist, width=bin_width, alpha=0.7, 
                                         color='blue', edgecolor='black', linewidth=0.5)
                        
                        # Mark the mode with a vertical line
                        self.ratio_ax.axvline(x=mode_value, color='red', linestyle='-', linewidth=2, 
                                             label=f'Mode: {mode_value:.3f}')
                        
                        # Show the current range being adjusted
                        self.ratio_ax.axvline(x=range_min, color='blue', linestyle='--', linewidth=1.5, 
                                             label=f'Min: {range_min:.3f}')
                        self.ratio_ax.axvline(x=range_max, color='blue', linestyle='--', linewidth=1.5, 
                                             label=f'Max: {range_max:.3f}')
                        
                        # Highlight the selected range
                        self.ratio_ax.axvspan(range_min, range_max, alpha=0.2, color='green',
                                            label=f'Range: {range_min:.3f}-{range_max:.3f}')
                        
                        # Set labels and title
                        self.ratio_ax.set_xlabel('Pulse Ratio', fontsize=10, fontweight='bold')
                        self.ratio_ax.set_ylabel('Frequency', fontsize=10, fontweight='bold')
                        self.ratio_ax.set_title('Distribution of Pulse Ratios', fontsize=12, fontweight='bold')
                        
                        # Restore previous axis limits to maintain zoom level
                        self.ratio_ax.set_xlim(current_xlim)
                        if current_ylim is not None:
                            self.ratio_ax.set_ylim(current_ylim)
                        
                        # Set grid
                        self.ratio_ax.grid(True, linestyle='--', alpha=0.7)
                        
                        # Make tick labels more visible
                        self.ratio_ax.tick_params(axis='both', which='major', labelsize=9)
                        
                        # Add a legend
                        self.ratio_ax.legend(loc='upper right', fontsize=8)
                        
                        # Update both canvases
                        preview_canvas.draw()
                        self.ratio_canvas.draw()
                
                # Connect value changes to update preview
                left_variation_input.valueChanged.connect(update_preview)
                right_variation_input.valueChanged.connect(update_preview)
                
                # Initial update
                update_preview()
                
                layout.addLayout(variation_layout)
                layout.addWidget(preview_label)
                
                # Add buttons
                buttons = QHBoxLayout()
                ok_button = QPushButton("OK")
                ok_button.clicked.connect(dialog.accept)
                cancel_button = QPushButton("Cancel")
                cancel_button.clicked.connect(dialog.reject)
                buttons.addWidget(ok_button)
                buttons.addWidget(cancel_button)
                layout.addLayout(buttons)
                
                if dialog.exec_() == QDialog.Accepted:
                    # Get the custom variation values
                    left_variation = left_variation_input.value()
                    right_variation = right_variation_input.value()
                    range_min = max(0, mode_value - left_variation)
                    range_max = min(1, mode_value + right_variation)
                else:
                    # User cancelled the custom variation dialog, use recommended values
                    range_min = recommended_min
                    range_max = recommended_max
            else:
                # User accepted the recommended variation
                range_min = recommended_min
                range_max = recommended_max
            
            # Store the ratio range
            self.ratio_range = (range_min, range_max)
            
            # Clear previous range visualization if any
            self.update_ratio_histogram()
            
            # Highlight the selected range on the histogram
            self.ratio_ax.axvspan(range_min, range_max, alpha=0.3, color='green', 
                                  label=f'Selected: {range_min:.3f}-{range_max:.3f}')
            
            # Update the canvas
            self.ratio_canvas.draw()
            
            # Update status
            self.status_label.setText(f"Selected ratio range: {range_min:.3f} - {range_max:.3f}")
            
            # Update the table with pulse start/end times and range indicators
            self.update_table_with_ranges()
            
            # Switch to the CSV table tab
            self.tabs.setCurrentIndex(1)  # Assuming CSV table is tab index 1
        else:
            # User cancelled
            self.status_label.setText("Ratio mode selection cancelled")

    def update_table_with_ranges(self):
        # Update the table with pulse pattern classification
        if not hasattr(self, 'period_range') or not hasattr(self, 'ratio_range'):
            QMessageBox.warning(self, "Warning", "Period and ratio ranges must be set first")
            return
            
        period_min, period_max = self.period_range
        ratio_min, ratio_max = self.ratio_range
        
        # Make sure we have a table to update
        if not hasattr(self, 'table'):
            print("Table not found, trying to use csv_table instead")
            if hasattr(self, 'csv_table'):
                self.table = self.csv_table
            else:
                QMessageBox.warning(self, "Warning", "Table not found")
                return
        
        # Check if we need to add the Pattern column
        if self.table.columnCount() < 8:  # We need at least 8 columns with TimeB added
            current_count = self.table.columnCount()
            self.table.setColumnCount(8)
            
            # Set headers for new columns if needed
            headers = []
            for i in range(current_count):
                if self.table.horizontalHeaderItem(i):
                    headers.append(self.table.horizontalHeaderItem(i).text())
                else:
                    headers.append(f"Column {i+1}")
            
            # Add headers for additional columns
            while len(headers) < 8:
                if len(headers) == 6:
                    headers.append("ex & in")
                elif len(headers) == 7:
                    headers.append("Sequencing")
            
            self.table.setHorizontalHeaderLabels(headers)
        
        print("\n\nDEBUGGING PULSE CLASSIFICATION:")
        print(f"Period range: {period_min:.2f}-{period_max:.2f}, Ratio range: {ratio_min:.4f}-{ratio_max:.4f}")
        print(f"Total pulses: {len(self.periods)}, Table rows: {self.table.rowCount()}")
        
        # First pass: Mark all short pulses as 'ex' if they're within ranges
        for row in range(min(self.table.rowCount(), len(self.periods))):
            try:
                period = self.periods[row]
                ratio = period.get('ratio', 0)
                duration = period.get('duration', 0)
                
                # Check if within both ranges
                within_period_range = period_min <= duration <= period_max
                within_ratio_range = ratio_min <= ratio <= ratio_max
                
                # Mark short pulses
                if ratio < 0.5 and within_period_range and within_ratio_range:
                    # Valid short pulse
                    self.table.setItem(row, 6, QTableWidgetItem("ex"))  # ex & in column (index 6 with TimeB added)
                    period['is_valid_short'] = True
                    print(f"Row {row}: Ratio={ratio:.4f}, Duration={duration:.2f} - Marking as short pulse (ex)")
                else:
                    # Invalid or not a short pulse
                    self.table.setItem(row, 6, QTableWidgetItem("z"))
                    period['is_valid_short'] = False
                    print(f"Row {row}: Ratio={ratio:.4f}, Duration={duration:.2f} - Temporarily marking as z")
            except Exception as e:
                print(f"Error in first pass, row {row}: {str(e)}")
        
        # Second pass: Mark long pulses as 'in' if they follow a valid short pulse
        for row in range(min(self.table.rowCount(), len(self.periods))):
            try:
                period = self.periods[row]
                ratio = period.get('ratio', 0)
                duration = period.get('duration', 0)
                
                # Check if within both ranges
                within_period_range = period_min <= duration <= period_max
                within_ratio_range = ratio_min <= ratio <= ratio_max
                
                # Check if this is a potential long pulse
                if ratio >= 0.5 and row > 0:
                    # Check if previous pulse was a valid short pulse
                    prev_period = self.periods[row-1]
                    if prev_period.get('is_valid_short', False):
                        # Valid long pulse
                        self.table.setItem(row, 6, QTableWidgetItem("in"))  # ex & in column (index 6 with TimeB added)
                        print(f"Row {row}: Ratio={ratio:.4f} - Marking as long/internal pulse (in)")
                    else:
                        # Invalid long pulse (previous short pulse was invalid)
                        self.table.setItem(row, 6, QTableWidgetItem("z"))
                        print(f"Row {row}: Ratio={ratio:.4f} - Marking as z (previous short pulse invalid)")
            except Exception as e:
                print(f"Error in second pass, row {row}: {str(e)}")
        
        # Third pass: Set background colors
        for row in range(min(self.table.rowCount(), len(self.periods))):
            try:
                item = self.table.item(row, 6)  # ex & in column (index 6 with TimeB added)
                if item:
                    # Set background color based on classification
                    if item.text() in ["ex", "in"]:
                        # Valid pulse (white background)
                        background_color = QColor(255, 255, 255)
                    else:
                        # Invalid pulse (red background)
                        background_color = QColor(255, 200, 200)
                    
                    # Apply background color to all cells in the row
                    for col in range(self.table.columnCount()):
                        cell_item = self.table.item(row, col)
                        if cell_item:
                            cell_item.setBackground(QBrush(background_color))
            except Exception as e:
                print(f"Error in third pass, row {row}: {str(e)}")
        
        # Fourth pass: Add sequencing information
        in_sequence = False
        sequence_start = -1
        last_valid_row = -1
        last_valid_type = ""
        
        for row in range(min(self.table.rowCount(), len(self.periods))):
            try:
                item = self.table.item(row, 6)  # ex & in column (now index 6 with TimeB added)
                if item:
                    if item.text() == "ex" or item.text() == "in":
                        # Valid pulse
                        if not in_sequence and item.text() == "ex":
                            # Start of a new sequence
                            in_sequence = True
                            sequence_start = row
                            # Mark as beginning
                            self.table.setItem(row, 7, QTableWidgetItem("B"))
                        elif in_sequence:
                            # Continuing sequence
                            self.table.setItem(row, 7, QTableWidgetItem("|"))
                        
                        # Update last valid pulse info
                        last_valid_row = row
                        last_valid_type = item.text()
                    else:
                        # Invalid pulse (z)
                        if in_sequence:
                            # End of sequence
                            in_sequence = False
                            # Mark the last valid pulse as end
                            if last_valid_row >= 0:
                                end_type = "E" + last_valid_type  # Either "Eex" or "Ein"
                                self.table.setItem(last_valid_row, 7, QTableWidgetItem(end_type))
            except Exception as e:
                print(f"Error in sequencing pass, row {row}: {str(e)}")
        
        # Handle case where sequence continues to the end of the table
        if in_sequence and last_valid_row >= 0:
            end_type = "E" + last_valid_type  # Either "Eex" or "Ein"
            self.table.setItem(last_valid_row, 7, QTableWidgetItem(end_type))
        
        # Update status
        self.status_label.setText(f"Updated pulse pattern classification")
        
        # Resize columns to fit content
        self.table.resizeColumnsToContents()
        
        # Make columns 1.25x bigger
        for col in range(self.table.columnCount()):
            width = self.table.columnWidth(col)
            self.table.setColumnWidth(col, int(width * 1.25))
        
        # Update status
        self.status_label.setText(f"Updated table with pulse times and range indicators. Period range: {period_min:.2f}-{period_max:.2f} ms, Ratio range: {ratio_min:.3f}-{ratio_max:.3f}")
    
    def create_ratio_histogram_tab(self):
        # Create tab widget
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Create matplotlib figure
        self.ratio_fig = Figure(figsize=(10, 5), tight_layout=True)
        self.ratio_canvas = FigureCanvas(self.ratio_fig)
        self.ratio_ax = self.ratio_fig.add_subplot(111)
        
        # Plot ratio histogram
        self.update_ratio_histogram()
        
        # Add canvas to layout
        layout.addWidget(self.ratio_canvas)
        
        # Add navigation instructions
        nav_label = QLabel("Navigation: W/S - Zoom In/Out, A/D - Move Left/Right, K - Select Mode, L - Set Range")
        nav_label.setFont(QFont("Arial", 9))
        nav_label.setStyleSheet("color: #666666;")
        layout.addWidget(nav_label)
        
        # Add info label for clicked bars
        self.ratio_info_label = QLabel("Click on a bar to see details")
        self.ratio_info_label.setFont(QFont("Arial", 9))
        self.ratio_info_label.setStyleSheet("color: #666666; background-color: #f0f0f0; padding: 5px; border-radius: 3px;")
        layout.addWidget(self.ratio_info_label)
        
        # Connect key press event
        self.ratio_canvas.setFocusPolicy(Qt.StrongFocus)
        self.ratio_canvas.mpl_connect('key_press_event', self.on_ratio_histogram_key_press)
        self.ratio_canvas.mpl_connect('button_press_event', self.on_ratio_histogram_click)
        
        # Initialize view limits
        self.ratio_view_limits = {'xmin': 0, 'xmax': 1, 'zoom_factor': 1.0}
        
        # Add the tab to the tab widget
        self.tabs.addTab(tab, "Ratio Histogram")
    
    def on_ratio_histogram_click(self, event):
        # Handle click events on the ratio histogram
        if event.inaxes != self.ratio_ax:
            return
        
        # Get the x value (ratio) where the user clicked
        clicked_x = event.xdata
        
        # Find periods close to the clicked value
        if hasattr(self, 'periods') and self.periods and hasattr(self, 'period_range'):
            period_min, period_max = self.period_range
            # Filter periods that are within the selected period range
            filtered_periods = [p for p in self.periods if 'duration' in p and period_min <= p['duration'] <= period_max]
            ratios = [p['ratio'] for p in filtered_periods if 'ratio' in p]
            
            if ratios:
                # Find the closest ratio to the clicked value
                closest_idx = np.argmin(np.abs(np.array(ratios) - clicked_x))
                closest_ratio = ratios[closest_idx]
                
                # Get the corresponding period data
                period_data = filtered_periods[closest_idx]
                
                # Update the info label with details
                info_text = f"Period: {period_data.get('period', 'N/A')}, Duration: {period_data.get('duration', 'N/A'):.2f} ms, "
                info_text += f"Ratio: {period_data.get('ratio', 'N/A'):.4f}, "
                info_text += f"Time: {period_data.get('time', 'N/A'):.2f} ms"
                
                self.ratio_info_label.setText(info_text)
    
    def on_ratio_histogram_key_press(self, event):
        # Handle key press events for ratio histogram navigation
        if not hasattr(self, 'ratio_view_limits'):
            self.ratio_view_limits = {'xmin': 0, 'xmax': 1, 'zoom_factor': 1.0}
        
        # Get current x limits
        xmin, xmax = self.ratio_ax.get_xlim()
        view_width = xmax - xmin
        
        # Handle key presses
        if event.key == 'w':  # Zoom in
            zoom_factor = 0.8
            center = (xmin + xmax) / 2
            new_width = view_width * zoom_factor
            self.ratio_ax.set_xlim(center - new_width/2, center + new_width/2)
            self.ratio_view_limits['zoom_factor'] *= zoom_factor
        elif event.key == 's':  # Zoom out
            zoom_factor = 1.25
            center = (xmin + xmax) / 2
            new_width = view_width * zoom_factor
            self.ratio_ax.set_xlim(center - new_width/2, center + new_width/2)
            self.ratio_view_limits['zoom_factor'] *= zoom_factor
        elif event.key == 'a':  # Move left
            move_amount = view_width * 0.2
            self.ratio_ax.set_xlim(xmin - move_amount, xmax - move_amount)
        elif event.key == 'd':  # Move right
            move_amount = view_width * 0.2
            self.ratio_ax.set_xlim(xmin + move_amount, xmax + move_amount)
        elif event.key == 'k':  # Select mode peak and set range
            self.select_ratio_mode_range()
        elif event.key == 'l':  # Manually set x-axis range
            self.set_ratio_histogram_range()
        
        # Update the canvas
        self.ratio_canvas.draw()
        
        # Update status
        self.status_label.setText(f"Ratio view: {xmin:.3f} - {xmax:.3f}, Zoom: {1/self.ratio_view_limits['zoom_factor']:.1f}x")
        
    def select_ratio_mode_range(self):
        # Select mode peak and set range for ratio histogram
        if hasattr(self, 'ratio_hist_data') and self.ratio_hist_data:
            mode_value = self.ratio_hist_data['mode_value']
            left_bar = self.ratio_hist_data['left_bar']
            right_bar = self.ratio_hist_data['right_bar']
            recommended_min = self.ratio_hist_data['recommended_min']
            recommended_max = self.ratio_hist_data['recommended_max']
            
            # Create a simple dialog asking if the variation is OK
            message = f"Mode peak detected at {mode_value:.3f}\n"
            message += f"Closest bars at {left_bar:.3f} and {right_bar:.3f}\n"
            message += f"Recommended range: {recommended_min:.3f} - {recommended_max:.3f}\n\n"
            message += "Is this range acceptable?"
            
            reply = QMessageBox.question(self, "Confirm Ratio Range", message,
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            
            if reply == QMessageBox.No:
                # User wants to adjust the range
                dialog = QDialog(self)
                dialog.setWindowTitle("Adjust Ratio Range")
                # Make dialog non-modal so user can interact with main window
                dialog.setWindowModality(Qt.NonModal)
                layout = QVBoxLayout(dialog)
                
                # Add explanation label
                layout.addWidget(QLabel(f"Mode value: {mode_value:.3f}"))
                layout.addWidget(QLabel("Set custom values for left and right sides:"))
                
                # Add inputs for both left and right sides
                variation_layout = QGridLayout()
                
                # Left/Right label for clarity
                variation_layout.addWidget(QLabel("Left/Right"), 0, 0)
                
                # Left value (absolute)
                variation_layout.addWidget(QLabel("Left value:"), 1, 0)
                left_value_input = QDoubleSpinBox()
                left_value_input.setRange(0, mode_value)
                left_value_input.setValue(recommended_min)
                left_value_input.setDecimals(3)
                left_value_input.setSingleStep(0.01)
                variation_layout.addWidget(left_value_input, 1, 1)
                
                # Right value (absolute)
                variation_layout.addWidget(QLabel("Right value:"), 2, 0)
                right_value_input = QDoubleSpinBox()
                right_value_input.setRange(mode_value, 1)
                right_value_input.setValue(recommended_max)
                right_value_input.setDecimals(3)
                right_value_input.setSingleStep(0.01)
                variation_layout.addWidget(right_value_input, 2, 1)
                
                # Preview of resulting range
                preview_label = QLabel(f"Resulting range: {left_value_input.value():.3f} - {right_value_input.value():.3f}")
                
                # Create a figure for real-time histogram preview
                preview_figure = Figure(figsize=(6, 3), dpi=100)
                preview_canvas = FigureCanvas(preview_figure)
                preview_ax = preview_figure.add_subplot(111)
                
                # Function to update the preview histogram and the main histogram
                def update_preview_histogram():
                    # Clear the axis
                    preview_ax.clear()
                    
                    # Get current values
                    range_min = left_value_input.value()
                    range_max = right_value_input.value()
                    
                    # Update text preview
                    preview_label.setText(f"Resulting range: {range_min:.3f} - {range_max:.3f}")
                    
                    # Also update the main histogram in real-time
                    # Store current axis limits before clearing
                    current_xlim = self.ratio_ax.get_xlim() if hasattr(self.ratio_ax, 'get_xlim') else (0, 1)
                    current_ylim = self.ratio_ax.get_ylim() if hasattr(self.ratio_ax, 'get_ylim') else None
                    
                    # Clear the main ratio histogram
                    self.ratio_ax.clear()
                    
                    # Redraw the main histogram
                    if hasattr(self, 'periods') and self.periods and hasattr(self, 'period_range'):
                        period_min, period_max = self.period_range
                        filtered_periods = [p for p in self.periods if 'duration' in p and period_min <= p['duration'] <= period_max]
                        ratios = [p['ratio'] for p in filtered_periods if 'ratio' in p]
                        
                        if ratios and hasattr(self, 'ratio_bins'):
                            # Plot histogram with bars that touch each other (no gaps)
                            self.ratio_ax.hist(ratios, bins=self.ratio_bins, alpha=0.7, 
                                              color='blue', edgecolor='black', linewidth=0.5,
                                              align='left', rwidth=1.0)
                            
                            # Mark the mode with a vertical line
                            self.ratio_ax.axvline(x=mode_value, color='red', linestyle='-', linewidth=2, 
                                                 label=f'Mode: {mode_value:.3f}')
                            
                            # Show the current range being adjusted
                            self.ratio_ax.axvline(x=range_min, color='blue', linestyle='--', linewidth=1.5, 
                                                 label=f'Min: {range_min:.3f}')
                            self.ratio_ax.axvline(x=range_max, color='blue', linestyle='--', linewidth=1.5, 
                                                 label=f'Max: {range_max:.3f}')
                            
                            # Highlight the selected range
                            self.ratio_ax.axvspan(range_min, range_max, alpha=0.3, color='green',
                                                label=f'Range: {range_min:.3f}-{range_max:.3f}')
                            
                            # Set labels and title
                            self.ratio_ax.set_xlabel('Pulse Ratio', fontsize=10, fontweight='bold')
                            self.ratio_ax.set_ylabel('Frequency', fontsize=10, fontweight='bold')
                            self.ratio_ax.set_title('Distribution of Pulse Ratios', fontsize=12, fontweight='bold')
                            
                            # Set grid
                            self.ratio_ax.grid(True, linestyle='--', alpha=0.7)
                            
                            # Make tick labels more visible
                            self.ratio_ax.tick_params(axis='both', which='major', labelsize=9)
                            
                            # Add a legend
                            self.ratio_ax.legend(loc='upper right', fontsize=8)
                            
                            # Restore previous axis limits to maintain zoom level
                            self.ratio_ax.set_xlim(current_xlim)
                            if current_ylim is not None:
                                self.ratio_ax.set_ylim(current_ylim)
                            
                            # Update the main canvas
                            self.ratio_canvas.draw()
                    
                    # Plot the histogram using stored data
                    if hasattr(self, 'ratio_hist_data'):
                        hist_data = self.ratio_hist_data
                        bins = hist_data['bin_edges']
                        hist = hist_data['hist']
                        
                        # Plot the histogram
                        bin_centers = [(bins[i] + bins[i+1])/2 for i in range(len(bins)-1)]
                        bin_width = bins[1] - bins[0]
                        preview_ax.bar(bin_centers, hist, width=bin_width, alpha=0.7, 
                                      color='blue', edgecolor='black', linewidth=0.5)
                        
                        # Mark the mode
                        preview_ax.axvline(x=mode_value, color='red', linestyle='-', linewidth=2)
                        
                        # Highlight the selected range
                        preview_ax.axvspan(range_min, range_max, alpha=0.3, color='blue')
                        
                        # Set labels and limits
                        preview_ax.set_xlabel('Pulse Ratio')
                        preview_ax.set_ylabel('Frequency')
                        preview_ax.set_xlim(0, 1)  # Standard x-axis limits for ratio histogram
                        preview_ax.grid(True, linestyle='--', alpha=0.7)
                    
                    # Redraw the canvas
                    preview_canvas.draw()
                
                # Connect value changes to update preview
                left_value_input.valueChanged.connect(update_preview_histogram)
                right_value_input.valueChanged.connect(update_preview_histogram)
                
                # Initial update
                update_preview_histogram()
                
                layout.addLayout(variation_layout)
                layout.addWidget(preview_label)
                
                # Add buttons
                buttons = QHBoxLayout()
                ok_button = QPushButton("OK")
                ok_button.clicked.connect(dialog.accept)
                cancel_button = QPushButton("Cancel")
                cancel_button.clicked.connect(dialog.reject)
                buttons.addWidget(ok_button)
                buttons.addWidget(cancel_button)
                layout.addLayout(buttons)
                
                if dialog.exec_() == QDialog.Accepted:
                    # Get the custom values directly
                    range_min = left_value_input.value()
                    range_max = right_value_input.value()
                else:
                    # User cancelled the custom variation dialog, use recommended values
                    range_min = recommended_min
                    range_max = recommended_max
            else:
                # User accepted the recommended variation
                range_min = recommended_min
                range_max = recommended_max
            
            # Store the ratio range
            self.ratio_range = (range_min, range_max)
            
            # Clear previous range visualization if any
            self.update_ratio_histogram()
            
            # Highlight the selected range on the histogram
            self.ratio_ax.axvspan(range_min, range_max, alpha=0.3, color='green', 
                                   label=f'Selected: {range_min:.3f}-{range_max:.3f}')
            
            # Update the canvas
            self.ratio_canvas.draw()
            
            # Update status
            self.status_label.setText(f"Selected ratio range: {range_min:.3f} - {range_max:.3f}")
            
            # Update the table with the selected ranges
            self.update_table_with_ranges()
    
    def set_ratio_histogram_range(self):
        # Allow user to manually set the x-axis range for ratio histogram
        current_min, current_max = self.ratio_ax.get_xlim()
        
        # Create dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Set Ratio Histogram X-Axis Range")
        layout = QVBoxLayout(dialog)
        
        # Add explanation
        layout.addWidget(QLabel("Set the x-axis range for the ratio histogram:"))
        
        # Add min/max inputs
        range_layout = QGridLayout()
        range_layout.addWidget(QLabel("Minimum:"), 0, 0)
        min_input = QDoubleSpinBox()
        min_input.setRange(0, 1)
        min_input.setValue(current_min)
        min_input.setDecimals(3)
        min_input.setSingleStep(0.01)
        range_layout.addWidget(min_input, 0, 1)
        
        range_layout.addWidget(QLabel("Maximum:"), 1, 0)
        max_input = QDoubleSpinBox()
        max_input.setRange(0, 1)
        max_input.setValue(current_max)
        max_input.setDecimals(3)
        max_input.setSingleStep(0.01)
        range_layout.addWidget(max_input, 1, 1)
        
        layout.addLayout(range_layout)
        
        # Add buttons
        buttons = QHBoxLayout()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(dialog.accept)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(dialog.reject)
        buttons.addWidget(ok_button)
        buttons.addWidget(cancel_button)
        layout.addLayout(buttons)
        
        if dialog.exec_() == QDialog.Accepted:
            # Get the custom range values
            new_min = min_input.value()
            new_max = max_input.value()
            
            # Validate range
            if new_min >= new_max:
                QMessageBox.warning(self, "Invalid Range", "Minimum must be less than maximum.")
                return
            
            # Set the new range
            self.ratio_ax.set_xlim(new_min, new_max)
            self.ratio_canvas.draw()
            
            # Update status
            self.status_label.setText(f"Ratio view range set to: {new_min:.3f} - {new_max:.3f}")
    
    def handle_copy_column(self):
        """Handle the Copy column functionality when user presses C key"""
        # Check if we have a table
        if not hasattr(self, 'table'):
            QMessageBox.warning(self, "Warning", "Table not found")
            return
        
        # Create a dialog for user settings
        settings_dialog = QDialog(self)
        settings_dialog.setWindowTitle("Copy Settings")
        settings_dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout(settings_dialog)
        
        # Add input fields with current values
        form_layout = QFormLayout()
        
        # Maximum error duration input
        max_error_input = QDoubleSpinBox()
        max_error_input.setRange(0.1, 1000.0)
        max_error_input.setSingleStep(1.0)
        max_error_input.setDecimals(1)
        max_error_input.setValue(self.max_error_duration_ms)
        form_layout.addRow("Maximum error duration (ms):", max_error_input)
        
        # Backward offset input
        backward_input = QDoubleSpinBox()
        backward_input.setRange(0, 1000.0)
        backward_input.setSingleStep(10.0)
        backward_input.setDecimals(1)
        backward_input.setValue(self.copy_backward_ms)
        form_layout.addRow("Time before end marker (ms):", backward_input)
        
        # Forward offset input
        forward_input = QDoubleSpinBox()
        forward_input.setRange(0, 1000.0)
        forward_input.setSingleStep(10.0)
        forward_input.setDecimals(1)
        forward_input.setValue(self.copy_forward_ms)
        form_layout.addRow("Time after begin marker (ms):", forward_input)
        
        layout.addLayout(form_layout)
        
        # Add buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(settings_dialog.accept)
        button_box.rejected.connect(settings_dialog.reject)
        layout.addWidget(button_box)
        
        # Show dialog and get result
        if settings_dialog.exec_() != QDialog.Accepted:
            return
        
        # Save the user's settings
        self.max_error_duration_ms = max_error_input.value()
        self.copy_backward_ms = backward_input.value()
        self.copy_forward_ms = forward_input.value()
        
        
        for row in range(self.table.rowCount()):
            # Clear the copy column
            self.table.setItem(row, 8, QTableWidgetItem(""))
        
        # Reset any class variables that might be storing copy information
        if hasattr(self, 'copy_sections'):
            del self.copy_sections
            
        # Reset the global copy counter to ensure we start from 1
        if hasattr(self, '_copy_counter'):
            del self._copy_counter
        self._copy_counter = 0  # Will be incremented to 1 for the first copy
        
        # Find error sequences (z values in the ex & in column)
        error_sequences = []
        current_error_seq = None
        
        # First pass: Find all rows with 'z' in the ex & in column
        for row in range(self.table.rowCount()):
            ex_in_item = self.table.item(row, 6)  # ex & in column (now index 6 with TimeB added)
            time_item = self.table.item(row, 2)   # Time column
            
            if not ex_in_item or not time_item or not time_item.text():
                continue
                
            try:
                time_value = float(time_item.text())
                ex_in_text = ex_in_item.text()
                
                # Check if this is an error row (z)
                if ex_in_text == "z":
                    # If we don't have an active error sequence, start one
                    if current_error_seq is None:
                        current_error_seq = {
                            "start_row": row,
                            "start_time": time_value,
                            "end_row": row,
                            "end_time": time_value,
                            "rows": [row]
                        }
                    else:
                        # Continue the current error sequence
                        current_error_seq["end_row"] = row
                        current_error_seq["end_time"] = time_value
                        current_error_seq["rows"].append(row)
                else:
                    # If we have an active error sequence and this is not an error row, end it
                    if current_error_seq is not None:
                        error_sequences.append(current_error_seq)
                        current_error_seq = None
            except ValueError:
                pass
        
        # Don't forget to add the last error sequence if it's still active
        if current_error_seq is not None:
            error_sequences.append(current_error_seq)
        
        # Filter error sequences by duration if needed
        if self.max_error_duration_ms > 0:
            filtered_sequences = []
            for seq in error_sequences:
                # Process each error sequence for copying
                # Find the end marker (E) before the error sequence
                end_marker_row = -1
                end_marker_time = None
                
                for row in range(seq["start_row"]-1, -1, -1):
                    seq_item = self.table.item(row, 7)  # Sequencing column (now index 7 with TimeB added)
                    time_item = self.table.item(row, 2)  # Time column
                    ex_in_item = self.table.item(row, 6)  # ex & in column (now index 6 with TimeB added)
                    
                    if (seq_item and time_item and time_item.text() and 
                        seq_item.text().startswith("E") and  # Ein or Eex
                        ex_in_item and ex_in_item.text() in ["ex", "in"]):
                        try:
                            end_marker_time = float(time_item.text())
                            break
                        except ValueError:
                            pass
                
                # Find the begin marker (B) after the error sequence
                begin_marker_row = -1
                begin_marker_time = None
                
                for row in range(seq["end_row"]+1, self.table.rowCount()):
                    seq_item = self.table.item(row, 7)  # Sequencing column (now index 7 with TimeB added)
                    time_item = self.table.item(row, 2)  # Time column
                    ex_in_item = self.table.item(row, 6)  # ex & in column (now index 6 with TimeB added)
                    
                    if (seq_item and time_item and time_item.text() and 
                        seq_item.text() == "B" and 
                        ex_in_item and ex_in_item.text() in ["ex", "in"]):
                        try:
                            begin_marker_time = float(time_item.text())
                            break
                        except ValueError:
                            pass
                
                # If we found both markers, check the duration between them
                if end_marker_time is not None and begin_marker_time is not None:
                    # Calculate duration between E and B markers
                    duration = begin_marker_time - end_marker_time
                    # Add to filtered sequences if duration is within max error duration
                    if duration <= self.max_error_duration_ms:
                        filtered_sequences.append(seq)
            error_sequences = filtered_sequences
        
        # Process each error sequence for copying
        for error_seq in error_sequences:
            # Find the end marker (E) before the error sequence
            end_marker_row = -1
            end_marker_time = None
            
            for row in range(error_seq["start_row"]-1, -1, -1):
                seq_item = self.table.item(row, 7)  # Sequencing column (now index 7 with TimeB added)
                time_item = self.table.item(row, 2)  # Time column
                ex_in_item = self.table.item(row, 6)  # ex & in column (now index 6 with TimeB added)
                
                if (seq_item and time_item and time_item.text() and 
                    seq_item.text().startswith("E") and  # Ein or Eex
                    ex_in_item and ex_in_item.text() in ["ex", "in"]):
                    try:
                        end_marker_row = row
                        end_marker_time = float(time_item.text())
                        break
                    except ValueError:
                        pass
            
            # Find the begin marker (B) after the error sequence
            begin_marker_row = -1
            begin_marker_time = None
            
            for row in range(error_seq["end_row"]+1, self.table.rowCount()):
                seq_item = self.table.item(row, 7)  # Sequencing column (now index 7 with TimeB added)
                time_item = self.table.item(row, 2)  # Time column
                ex_in_item = self.table.item(row, 6)  # ex & in column (now index 6 with TimeB added)
                
                if (seq_item and time_item and time_item.text() and 
                    seq_item.text() == "B" and 
                    ex_in_item and ex_in_item.text() in ["ex", "in"]):
                    try:
                        begin_marker_row = row
                        begin_marker_time = float(time_item.text())
                        break
                    except ValueError:
                        pass
            
            # If we found both markers, create a copy section
            if end_marker_time is not None and begin_marker_time is not None:
                # Increment the copy counter
                self._copy_counter += 1
                current_copy_number = self._copy_counter
                
                # Calculate copy range using user-defined offsets
                copy_start_time = end_marker_time - self.copy_backward_ms
                copy_end_time = begin_marker_time + self.copy_forward_ms
                
                # Find rows that fall within this time range
                for row in range(self.table.rowCount()):
                    time_item = self.table.item(row, 2)  # Time column
                    if time_item and time_item.text():
                        try:
                            time_value = float(time_item.text())
                            if copy_start_time <= time_value <= copy_end_time:
                                # Mark this row for copying and highlight in yellow
                                copy_item = QTableWidgetItem(f"Copy {current_copy_number}")
                                copy_item.setBackground(QBrush(QColor(255, 255, 0)))  # Yellow background
                                self.table.setItem(row, 8, copy_item)
                        except ValueError:
                            pass
        
        # Update status
        if self._copy_counter > 0:
            self.status_label.setText(f"Found {self._copy_counter} error sequences to copy")
        else:
            self.status_label.setText("No error sequences found matching the criteria")
    
    def save_waveform_files(self):
        """Save waveform segments as WAV files when user presses = key"""
        # Check if we have a table and waveform data
        if not hasattr(self, 'table'):
            QMessageBox.warning(self, "Warning", "Table not found")
            return
        
        if not hasattr(self, 'wav_data') or self.wav_data is None or not hasattr(self, 'sample_rate') or self.sample_rate is None:
            QMessageBox.warning(self, "Warning", "Waveform data not loaded")
            return
        
        # Prompt user for folder name
        folder_name, ok = QInputDialog.getText(
            self, "Save Waveform Files", 
            "Enter folder name for waveform files:")
        
        if not ok or not folder_name:
            return
        
        # Create directory if it doesn't exist
        if hasattr(self, 'wav_file_path') and self.wav_file_path:
            save_dir = os.path.join(os.path.dirname(self.wav_file_path), folder_name)
        elif hasattr(self, 'csv_file_path') and self.csv_file_path:
            save_dir = os.path.join(os.path.dirname(self.csv_file_path), folder_name)
        else:
            save_dir = os.path.join(os.getcwd(), folder_name)
        
        os.makedirs(save_dir, exist_ok=True)
        print(f"Created directory: {save_dir}")
        
        # Create DataFrame from table data for the total CSV
        headers = []
        for col in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(col)
            if header_item:
                headers.append(header_item.text())
            else:
                headers.append(f"Column {col+1}")
        
        data = []
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    row_data.append(item.text())
                else:
                    row_data.append("")
            data.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Save total CSV file
        total_file_path = os.path.join(save_dir, f"{folder_name}_TOTAL.csv")
        df.to_csv(total_file_path, index=False)
        
        # Find all copy sections
        copy_sections = {}
        for row in range(self.table.rowCount()):
            copy_item = self.table.item(row, 8)  # Copy column (index 8 with TimeB added)
            if copy_item and copy_item.text() and copy_item.text().startswith("Copy "):
                copy_num = copy_item.text().split(" ")[1]
                if copy_num not in copy_sections:
                    copy_sections[copy_num] = []
                copy_sections[copy_num].append(row)
        
        # Print detailed information about found copy sections
        print(f"Found {len(copy_sections)} copy sections in the table")
        for copy_num in sorted(copy_sections.keys(), key=int):
            print(f"Copy section {copy_num} has {len(copy_sections[copy_num])} rows")
        if not copy_sections:
            print("WARNING: No copy sections found in the table. Did you press 'C' to mark sections for copying?")
            QMessageBox.warning(self, "No Copy Sections", "No copy sections found in the table. Press 'C' to mark sections for copying first.")
            # Continue anyway to at least save the CSV and Excel files
        
        # Create a combined waveform with all segments
        combined_waveform = np.array([], dtype=np.int16)
        one_second_samples = 1 * self.sample_rate  # Exactly one second of samples
        
        # Process each copy section and add to the combined waveform
        wav_files_saved = 0
        waveform_segments = []
        
        for copy_num, rows in copy_sections.items():
            if rows:
                # Get time values for this section
                time_values = []
                for row in rows:
                    time_item = self.table.item(row, 2)  # Time column
                    if time_item and time_item.text():
                        try:
                            time_values.append(float(time_item.text()))
                        except ValueError:
                            pass
                
                if time_values:
                    # Calculate start and end times in milliseconds
                    min_time = min(time_values)
                    max_time = max(time_values)
                    
                    # Calculate the range to extract (E - 200ms to B + 300ms)
                    # Find the last valid pulse before this error sequence
                    start_row = min(rows)
                    last_valid_time = None
                    for row in range(start_row-1, -1, -1):
                        item = self.table.item(row, 6)  # ex & in column (now index 6 with TimeB added)
                        time_item = self.table.item(row, 2)  # Time column
                        if item and time_item and item.text() in ["ex", "in"] and time_item.text():
                            try:
                                last_valid_time = float(time_item.text())
                                break
                            except ValueError:
                                pass
                    
                    # Find the first valid pulse after this error sequence
                    end_row = max(rows)
                    next_valid_time = None
                    for row in range(end_row+1, self.table.rowCount()):
                        item = self.table.item(row, 6)  # ex & in column (now index 6 with TimeB added)
                        time_item = self.table.item(row, 2)  # Time column
                        if item and time_item and item.text() in ["ex", "in"] and time_item.text():
                            try:
                                next_valid_time = float(time_item.text())
                                break
                            except ValueError:
                                pass
                    
                    # If we found valid pulses before and after
                    if last_valid_time is not None and next_valid_time is not None:
                        # Calculate the waveform segment to extract using user-defined offsets
                        extract_start_ms = last_valid_time - self.copy_backward_ms  # User-defined ms before last valid pulse
                        extract_end_ms = next_valid_time + self.copy_forward_ms      # User-defined ms after next valid pulse
                        
                        # Print detailed information about this segment
                        print(f"Copy {copy_num}: Last valid time: {last_valid_time:.2f}ms, Next valid time: {next_valid_time:.2f}ms")
                        print(f"Copy {copy_num}: Extract range: {extract_start_ms:.2f}ms to {extract_end_ms:.2f}ms (duration: {extract_end_ms-extract_start_ms:.2f}ms)")
                        
                        # Convert milliseconds to sample indices
                        start_sample = max(0, int((extract_start_ms / 1000) * self.sample_rate))
                        end_sample = min(len(self.wav_data), int((extract_end_ms / 1000) * self.sample_rate))
                        
                        # Extract the waveform segment
                        waveform_segment = self.wav_data[start_sample:end_sample]
                        
                        # Store the segment for later combining
                        waveform_segments.append((copy_num, waveform_segment))
                        wav_files_saved += 1
                    else:
                        # Print debug info about why we couldn't find valid pulses
                        print(f"WARNING: Copy {copy_num} - Could not find valid pulses before and/or after error sequence")
                        print(f"  Last valid time: {last_valid_time}, Next valid time: {next_valid_time}")
                        
                        # Even if we don't have both markers, try to create a segment with what we have
                        if last_valid_time is not None or next_valid_time is not None:
                            # If we only have one marker, use a fixed duration for the other side
                            if last_valid_time is not None and next_valid_time is None:
                                # We have the start but not the end, use a fixed duration (e.g., 500ms after start)
                                extract_start_ms = last_valid_time - self.copy_backward_ms
                                extract_end_ms = last_valid_time + 500  # 500ms after last valid time as a fallback
                                print(f"  Using fallback end time: {extract_end_ms:.2f}ms")
                            elif next_valid_time is not None and last_valid_time is None:
                                # We have the end but not the start, use a fixed duration (e.g., 500ms before end)
                                extract_start_ms = next_valid_time - 500  # 500ms before next valid time as a fallback
                                extract_end_ms = next_valid_time + self.copy_forward_ms
                                print(f"  Using fallback start time: {extract_start_ms:.2f}ms")
                            
                            # Ensure start time is not negative
                            extract_start_ms = max(0, extract_start_ms)
                            
                            # Convert milliseconds to sample indices
                            start_sample = max(0, int((extract_start_ms / 1000) * self.sample_rate))
                            end_sample = min(len(self.wav_data), int((extract_end_ms / 1000) * self.sample_rate))
                            
                            # Extract the waveform segment
                            waveform_segment = self.wav_data[start_sample:end_sample]
                            
                            # Store the segment for later combining
                            waveform_segments.append((copy_num, waveform_segment))
                            wav_files_saved += 1
                            print(f"  Created fallback segment for Copy {copy_num} from {extract_start_ms:.2f}ms to {extract_end_ms:.2f}ms")
                        else:
                            print(f"  Cannot create segment for Copy {copy_num} - no valid markers found")
                            # Create a short empty segment as a placeholder so we don't lose the copy number
                            empty_segment = np.zeros(int(0.5 * self.sample_rate), dtype=np.float64)  # 0.5 second of silence
                            waveform_segments.append((copy_num, empty_segment))
                            wav_files_saved += 1
                            print(f"  Created empty placeholder segment for Copy {copy_num}")
                else:
                    print(f"WARNING: Copy {copy_num} - No time values found in the rows")
                    # Create a short empty segment as a placeholder
                    empty_segment = np.zeros(int(0.5 * self.sample_rate), dtype=np.float64)  # 0.5 second of silence
                    waveform_segments.append((copy_num, empty_segment))
                    wav_files_saved += 1
                    print(f"  Created empty placeholder segment for Copy {copy_num}")
            else:
                print(f"WARNING: Copy {copy_num} - No rows found")
                # Create a short empty segment as a placeholder
                empty_segment = np.zeros(int(0.5 * self.sample_rate), dtype=np.float64)  # 0.5 second of silence
                waveform_segments.append((copy_num, empty_segment))
                wav_files_saved += 1
                print(f"  Created empty placeholder segment for Copy {copy_num}")
        
        # If we have waveform segments, combine them into a single file
        if waveform_segments:
            # Sort segments by copy number
            waveform_segments.sort(key=lambda x: int(x[0]))
            
            # Calculate the total length needed for all segments with exactly one second per segment
            total_length = one_second_samples * len(waveform_segments)
            combined_waveform = np.zeros(total_length, dtype=np.int16)
            
            # Add each segment to the combined waveform at one-second intervals
            for i, (copy_num, segment) in enumerate(waveform_segments):
                # Scale to int16 range
                if np.max(np.abs(segment)) > 0:
                    segment = segment / np.max(np.abs(segment)) * 32767
                
                # Convert to int16
                segment_int16 = segment.astype(np.int16)
                
                # Calculate the start position for this segment
                start_pos = i * one_second_samples
                
                # Make sure we preserve the entire segment without truncation
                segment_length = len(segment_int16)
                
                # If the segment is longer than one second, we need to extend the combined waveform
                if segment_length > one_second_samples:
                    # Create a new combined waveform with enough space
                    new_total_length = total_length + (segment_length - one_second_samples)
                    new_combined_waveform = np.zeros(new_total_length, dtype=np.int16)
                    
                    # Copy existing data
                    new_combined_waveform[:start_pos] = combined_waveform[:start_pos]
                    
                    # Adjust positions for subsequent segments
                    if i < len(waveform_segments) - 1:
                        # Calculate how much we need to shift subsequent segments
                        shift_amount = segment_length - one_second_samples
                        # Copy the rest of the data with the shift
                        new_combined_waveform[start_pos + segment_length:] = combined_waveform[start_pos + one_second_samples:]
                    
                    # Update the combined waveform
                    combined_waveform = new_combined_waveform
                    total_length = new_total_length
                
                # Add the segment to the combined waveform (full length, no truncation)
                combined_waveform[start_pos:start_pos + segment_length] = segment_int16
            
            # Save the combined waveform as a WAV file
            combined_wav_path = os.path.join(save_dir, f"{folder_name}_combined.wav")
            print(f"Saving combined WAV file: {combined_wav_path}")
            try:
                scipy.io.wavfile.write(combined_wav_path, self.sample_rate, combined_waveform)
                print(f"Successfully saved combined WAV file: {combined_wav_path}")
            except Exception as e:
                print(f"Error saving combined WAV file: {str(e)}")
                QMessageBox.warning(self, "Error", f"Error saving combined WAV file: {str(e)}")
                import traceback
                traceback.print_exc()
        
        # Create Excel file with formatted data if openpyxl is available
        excel_saved = False
        if EXCEL_EXPORT_AVAILABLE:  # Check if openpyxl is available
            try:
                # Create Excel workbook
                excel_file_path = os.path.join(save_dir, f"{folder_name}_exceltotal.xlsx")
                print(f"Creating Excel file at: {excel_file_path}")
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Total Data"
                
                # Add data from DataFrame to Excel
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Format header row
                        if r_idx == 1:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Highlight cells with "Copy" in column 8
                        if c_idx == 8 and r_idx > 1 and value and value.startswith("Copy"):
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.25
                    sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Save Excel file
                workbook.save(excel_file_path)
                excel_saved = True
            except Exception as e:
                print(f"Error creating Excel file: {str(e)}")
                excel_saved = False
        
        # Update status
        if excel_saved:
            self.status_label.setText(f"Saved 1 CSV file, {wav_files_saved} WAV files, and Excel file to {save_dir}")
            QMessageBox.information(self, "Files Saved", f"Saved 1 CSV file, {wav_files_saved} WAV files, and Excel file to {save_dir}")
        else:
            self.status_label.setText(f"Saved 1 CSV file and {wav_files_saved} WAV files to {save_dir}")
            QMessageBox.information(self, "Files Saved", f"Saved 1 CSV file and {wav_files_saved} WAV files to {save_dir}. Excel export failed.")
    
    def update_ratio_histogram(self):
        # Check if we have the period range set
        if not hasattr(self, 'period_range'):
            return
        
        # Clear the plot if ratio_ax exists
        if hasattr(self, 'ratio_ax'):
            self.ratio_ax.clear()
        else:
            return
        
        # Plot ratio histogram
        if hasattr(self, 'periods') and self.periods:
            period_min, period_max = self.period_range
            
            # Filter periods that are within the selected period range
            filtered_periods = [p for p in self.periods if 'duration' in p and period_min <= p['duration'] <= period_max]
            ratios = [p['ratio'] for p in filtered_periods if 'ratio' in p]
            
            if ratios:
                # Set fixed view limits to 0-1 as requested
                self.ratio_ax.set_xlim(0, 1)
                
                # Create histogram with evenly spaced bins to avoid missing bars
                # Using 100 bins from 0 to 1 to ensure fine granularity
                hist, bin_edges = np.histogram(ratios, bins=np.linspace(0, 1, 101))
                
                # Compute bin centers for plotting
                bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
                bin_width = bin_edges[1] - bin_edges[0]
                
                # Plot using bar to control exact bin locations
                self.ratio_ax.bar(bin_centers, hist, width=bin_width,
                                 align='center', color='blue', edgecolor='black', alpha=0.7)
                
                # Store the bin information for later use
                self.ratio_bins = bin_edges
                
                # Find the actual most common value in the dataset
                unique_values, counts = np.unique(ratios, return_counts=True)
                mode_idx = np.argmax(counts)
                mode_value = unique_values[mode_idx]
                self.ratio_mode = mode_value
                
                # Find the bin that contains the mode value
                mode_bin_idx = np.digitize(mode_value, bin_edges) - 1
                mode_bin_center = bin_centers[mode_bin_idx]
                
                # Print for debugging
                print(f"Most common ratio value (actual mode): {mode_value:.3f}")
                print(f"Mode bin center: {mode_bin_center:.3f}")
                print(f"Bin edges of mode bin: {bin_edges[mode_bin_idx]:.3f} - {bin_edges[mode_bin_idx + 1]:.3f}")
                
                # Find the bar 2 positions to the left of the mode bar (if it exists)
                if mode_bin_idx > 1:
                    left_bar = bin_centers[mode_bin_idx - 2]
                elif mode_bin_idx > 0:
                    left_bar = bin_centers[mode_bin_idx - 1]
                else:
                    left_bar = max(0, mode_bin_center - 2 * bin_width)
                
                # Find the bar 2 positions to the right of the mode bar (if it exists)
                if mode_bin_idx < len(bin_centers) - 2:
                    right_bar = bin_centers[mode_bin_idx + 2]
                elif mode_bin_idx < len(bin_centers) - 1:
                    right_bar = bin_centers[mode_bin_idx + 1]
                else:
                    right_bar = min(1, mode_bin_center + 2 * bin_width)
                
                # Calculate the recommended range based on exactly two bars in front and two behind
                # The range should be from the left bar to the right bar (inclusive)
                recommended_min = left_bar
                recommended_max = right_bar
                
                # Store histogram data for later use
                self.ratio_hist_data = {
                    'hist': hist, 
                    'bin_edges': bin_edges, 
                    'bin_centers': bin_centers,
                    'mode_bin_idx': mode_bin_idx, 
                    'mode_value': mode_value,
                    'left_bar': left_bar,
                    'right_bar': right_bar,
                    'recommended_min': recommended_min,
                    'recommended_max': recommended_max
                }
                
                # Mark the mode with a vertical line
                self.ratio_ax.axvline(x=mode_value, color='red', linestyle='-', linewidth=2, 
                                       label=f'Mode: {mode_value:.3f}')
                
                # Mark the closest bars with more prominent lines
                self.ratio_ax.axvline(x=left_bar, color='blue', linestyle='-', linewidth=2, 
                                       label=f'Left: {left_bar:.3f}')
                self.ratio_ax.axvline(x=right_bar, color='blue', linestyle='-', linewidth=2, 
                                       label=f'Right: {right_bar:.3f}')
                
                # Show recommended range - exactly between the left and right bars
                self.ratio_ax.axvspan(left_bar, right_bar, alpha=0.2, color='green',
                                      label=f'Range: {left_bar:.3f}-{right_bar:.3f}')
                
                # Add a legend
                self.ratio_ax.legend(loc='upper right', fontsize=8)
        
        # Set labels and title
        self.ratio_ax.set_xlabel('Pulse Ratio', fontsize=10, fontweight='bold')
        self.ratio_ax.set_ylabel('Frequency', fontsize=10, fontweight='bold')
        self.ratio_ax.set_title('Distribution of Pulse Ratios', fontsize=12, fontweight='bold')
        
        # Set grid
        self.ratio_ax.grid(True, linestyle='--', alpha=0.7)
        
        # Make tick labels more visible
        self.ratio_ax.tick_params(axis='both', which='major', labelsize=9)
        
        # Update canvas
        self.ratio_canvas.draw()
    
    def update_period_histogram(self):
        # Clear the plot
        self.period_ax.clear()
        
        # Plot period histogram
        if hasattr(self, 'periods') and self.periods:
            durations = [p['duration'] for p in self.periods if 'duration' in p]
            if durations:
                # Set fixed view limits to 0-50ms as requested
                self.period_ax.set_xlim(0, 50)
                
                # Calculate bin width based on sampling rate to ensure no gaps
                # Find the minimum difference between consecutive sorted durations
                sorted_durations = sorted(durations)
                diffs = [sorted_durations[i+1] - sorted_durations[i] for i in range(len(sorted_durations)-1)]
                if diffs:
                    min_diff = min([d for d in diffs if d > 0], default=0.1)  # Minimum non-zero difference
                    # Use exactly the minimum difference as bin width to ensure no gaps
                    bin_width = min_diff
                else:
                    bin_width = 0.1  # Default if we can't determine from data
                
                # Create bins with appropriate granularity from 0 to 50ms
                bins = np.arange(0, 50 + bin_width, bin_width)
                
                # Store the bin information for later use
                self.period_bins = bins
                
                # First get the histogram data for visualization
                hist, bin_edges = np.histogram(durations, bins=bins)
                
                # Plot histogram with bars that touch each other (no gaps)
                n, bins_out, patches = self.period_ax.hist(durations, bins=bins, alpha=0.7, 
                                                 color='green', edgecolor='black', linewidth=0.5,
                                                 align='left', rwidth=1.0)  # rwidth=1.0 ensures bars touch
                
                # Find the actual most common value in the dataset
                unique_values, counts = np.unique(durations, return_counts=True)
                mode_idx = np.argmax(counts)
                mode_value = unique_values[mode_idx]
                self.period_mode = mode_value
                
                # Find the bin that contains the most data points (the tallest bar)
                tallest_bin_idx = np.argmax(hist)
                tallest_bin_center = (bin_edges[tallest_bin_idx] + bin_edges[tallest_bin_idx + 1]) / 2
                
                # Print for debugging
                print(f"Most common value (actual mode): {mode_value:.2f} ms")
                print(f"Tallest bar center: {tallest_bin_center:.2f} ms")
                print(f"Bin edges of tallest bar: {bin_edges[tallest_bin_idx]:.2f} - {bin_edges[tallest_bin_idx + 1]:.2f} ms")
                
                # Find the bin centers for all bins
                bin_centers = [(bin_edges[i] + bin_edges[i+1])/2 for i in range(len(bin_edges)-1)]
                
                # Find the bar 2 positions to the left of the tallest bar (if it exists)
                if tallest_bin_idx > 1:
                    left_bar = bin_centers[tallest_bin_idx - 2]
                elif tallest_bin_idx > 0:
                    left_bar = bin_centers[tallest_bin_idx - 1]
                else:
                    left_bar = max(0, mode_value - 2 * bin_width)
                
                # Find the bar 2 positions to the right of the tallest bar (if it exists)
                if tallest_bin_idx < len(bin_centers) - 2:
                    right_bar = bin_centers[tallest_bin_idx + 2]
                elif tallest_bin_idx < len(bin_centers) - 1:
                    right_bar = bin_centers[tallest_bin_idx + 1]
                else:
                    right_bar = min(50, mode_value + 2 * bin_width)
                
                # Calculate the recommended range based on exactly two bars in front and two behind
                # The range is simply from the left bar to the right bar
                recommended_min = max(0, left_bar)
                recommended_max = min(50, right_bar)
                
                # Store histogram data for later use
                self.period_hist_data = {
                    'hist': hist, 
                    'bin_edges': bin_edges, 
                    'tallest_bin_idx': tallest_bin_idx, 
                    'mode_value': mode_value,
                    'left_bar': left_bar,
                    'right_bar': right_bar,
                    'recommended_min': recommended_min,
                    'recommended_max': recommended_max
                }
                
                # Mark the mode with a vertical line
                self.period_ax.axvline(x=mode_value, color='red', linestyle='-', linewidth=2, 
                                       label=f'Mode: {mode_value:.2f} ms')
                
                # Mark the closest bars
                self.period_ax.axvline(x=left_bar, color='green', linestyle='--', linewidth=1.5, 
                                       label=f'Left: {left_bar:.2f} ms')
                self.period_ax.axvline(x=right_bar, color='green', linestyle='--', linewidth=1.5, 
                                       label=f'Right: {right_bar:.2f} ms')
                
                # Show recommended range
                self.period_ax.axvspan(recommended_min, recommended_max, alpha=0.2, color='blue',
                                      label=f'Range: {recommended_min:.2f}-{recommended_max:.2f} ms')
                
                # Add a legend
                self.period_ax.legend(loc='upper right', fontsize=8)
        
        # Set labels and title
        self.period_ax.set_xlabel('Period Duration (ms)', fontsize=10, fontweight='bold')
        self.period_ax.set_ylabel('Frequency', fontsize=10, fontweight='bold')
        self.period_ax.set_title('Distribution of Period Durations', fontsize=12, fontweight='bold')
        
        # Set grid
        self.period_ax.grid(True, linestyle='--', alpha=0.7)
        
        # Make tick labels more visible
        self.period_ax.tick_params(axis='both', which='major', labelsize=9)
        
        # Update canvas
        self.period_canvas.draw()
    
    # This is a duplicate method that was causing errors - removed
    
        
        
        
    def set_pulse_pattern_variation(self):
        try:
            print("Starting set_pulse_pattern_variation")
            # Calculate recommended variation based on the dataset
            recommended_variation = 0.05  # Default
            
            if hasattr(self, 'periods') and self.periods:
                ratios = [p.get('ratio', 0.5) for p in self.periods if 'ratio' in p]
                if ratios:
                    # Calculate standard deviation around 0.5
                    ratios_near_half = [r for r in ratios if 0.4 <= r <= 0.6]
                    if ratios_near_half:
                        std_dev = np.std(ratios_near_half)
                        recommended_variation = max(0.02, min(0.1, std_dev * 2))  # Reasonable bounds
            
            print(f"Recommended variation: {recommended_variation}")
            
            # Create a simple input dialog instead of a complex one
            variation, ok = QInputDialog.getDouble(
                self, "Set Pulse Pattern Variation",
                "Enter variation around 0.5 for single pulse pattern classification:\n"
                "Ratios within 0.5 ± variation will be classified as single pulse (SP),\n"
                "otherwise they will be classified as double pulse (DP).\n"
                f"Recommended: {recommended_variation:.3f}",
                recommended_variation, 0.001, 0.2, 3
            )
            
            if ok:
                print(f"Selected variation: {variation}")
                # Update variation
                self.pulse_pattern_variation = variation
                # Update pulse patterns in the table
                self.update_pulse_patterns()
                # Update status
                self.status_label.setText(f"Pulse pattern variation set to ±{self.pulse_pattern_variation:.3f}")
            else:
                print("Dialog cancelled")
        except Exception as e:
            print(f"Error in set_pulse_pattern_variation: {str(e)}")
            import traceback
            traceback.print_exc()
            self.status_label.setText(f"Error: {str(e)}")
    
    def determine_pulse_pattern(self, ratio):
        # Determine pulse pattern based on ratio and variation
        if 0.5 - self.pulse_pattern_variation <= ratio <= 0.5 + self.pulse_pattern_variation:
            return "SP"  # Single Pulse
        else:
            return "DP"  # Double Pulse
    
    def update_pulse_patterns(self):
        # Add a Pulse Pattern column to the table if it doesn't exist
        if self.table.columnCount() < 6:
            self.table.setColumnCount(6)
            headers = [self.table.horizontalHeaderItem(i).text() for i in range(5)]
            headers.insert(1, "Pulse Pattern")  # Add after Period column
            self.table.setHorizontalHeaderLabels(headers)
        
        # Update the pulse pattern for each row
        for row in range(self.table.rowCount()):
            # Get the ratio from the table
            ratio_item = self.table.item(row, 3)  # Pulse Ratio column
            if ratio_item and ratio_item.text():
                try:
                    ratio = float(ratio_item.text())
                    pattern = self.determine_pulse_pattern(ratio)
                    
                    # Create or update the pattern item
                    pattern_item = self.table.item(row, 1)
                    if not pattern_item:
                        pattern_item = QTableWidgetItem(pattern)
                        self.table.setItem(row, 1, pattern_item)
                    else:
                        pattern_item.setText(pattern)
                    
                    # Set background color based on pattern
                    if pattern == "SP":
                        pattern_item.setBackground(QBrush(QColor(200, 255, 200)))  # Light green
                    else:  # DP
                        pattern_item.setBackground(QBrush(QColor(255, 200, 200)))  # Light red
                except ValueError:
                    pass
        
        # Resize columns to content
        self.table.resizeColumnsToContents()
    
        
    def search_for_deviations(self):
        try:
            print("Starting search_for_deviations")
            # Find the mode of the durations
            if not hasattr(self, 'periods') or not self.periods:
                print("No periods data available")
                self.status_label.setText("No period data available for analysis")
                return
            
            durations = [p.get('duration', 0) for p in self.periods if 'duration' in p]
            if not durations:
                print("No duration data available")
                self.status_label.setText("No duration data available for analysis")
                return
            
            print(f"Found {len(durations)} durations")
            
            # Calculate mode
            hist, bin_edges = np.histogram(durations, bins=np.arange(0, 100, 0.1))
            mode_idx = np.argmax(hist)
            mode_value = (bin_edges[mode_idx] + bin_edges[mode_idx + 1]) / 2
            print(f"Mode value: {mode_value}")
            
            # Calculate recommended threshold (1 standard deviation)
            std_dev = np.std(durations)
            recommended_threshold = max(1.0, min(5.0, std_dev))  # Reasonable bounds
            print(f"Std dev: {std_dev}, Recommended threshold: {recommended_threshold}")
            
            # Create a simple input dialog instead of a complex one
            threshold, ok = QInputDialog.getDouble(
                self, "Set Deviation Threshold",
                f"Enter threshold for deviations from mode ({mode_value:.2f} ms):\n"
                f"Recommended: {recommended_threshold:.1f} ms (1 std dev)",
                recommended_threshold, 0.1, 20.0, 1
            )
            
            if ok:
                print(f"Selected threshold: {threshold}")
                # Highlight deviations
                self.highlight_deviations(mode_value, threshold)
                # Update status
                self.status_label.setText(f"Highlighted deviations from mode ({mode_value:.2f} ms) with threshold ±{threshold:.1f} ms")
            else:
                print("Dialog cancelled")
        except Exception as e:
            print(f"Error in search_for_deviations: {str(e)}")
            import traceback
            traceback.print_exc()
            self.status_label.setText(f"Error: {str(e)}")
    
    def highlight_deviations(self, mode_value, threshold):
        # Reset all cell backgrounds
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    item.setBackground(QBrush(QColor(255, 255, 255)))  # White background
        
        # Highlight rows with durations that deviate from the mode
        for row in range(self.table.rowCount()):
            duration_item = self.table.item(row, 1)  # Duration column
            if duration_item and duration_item.text():
                try:
                    duration = float(duration_item.text())
                    if abs(duration - mode_value) > threshold:
                        # This duration deviates from the mode
                        for col in range(self.table.columnCount()):
                            item = self.table.item(row, col)
                            if item:
                                # Light orange background for deviations
                                item.setBackground(QBrush(QColor(255, 220, 180)))
                except ValueError:
                    pass
        
        # Scroll to the first highlighted row
        for row in range(self.table.rowCount()):
            duration_item = self.table.item(row, 1)  # Duration column
            if duration_item and duration_item.text():
                try:
                    duration = float(duration_item.text())
                    if abs(duration - mode_value) > threshold:
                        self.table.scrollToItem(duration_item)
                        break
                except ValueError:
                    pass


# Initialize the application
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = KatydidAnalyzer2()
    window.show()
    sys.exit(app.exec_())